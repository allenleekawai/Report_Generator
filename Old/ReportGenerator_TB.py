# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import shutil
import numpy as np

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle

import time
from datetime import datetime

import tkinter as tk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

from functools import partial
from tenacity import retry, stop_after_delay, wait_fixed, retry_if_exception_type

execute = True

while True:

    totalStartTime = time.time() # 紀錄開始時間

    # ================================================================================ #
    #                                                                                  #
    #                             宣告 Function 及建立 GUI                              #
    #                                                                                  #
    # ================================================================================ #

    if True:

        if True: # 宣告 Function

            # 宣告 "選擇 Limit File" Function
            if True:
                def selectLimitFilePath1():
                    limitFilePath_1 = askopenfilename()
                    limitFilePath1.set(limitFilePath_1)

                def selectLimitFilePath2():
                    limitFilePath_2 = askopenfilename()
                    limitFilePath2.set(limitFilePath_2)

                def selectLimitFilePath3():
                    limitFilePath_3 = askopenfilename()
                    limitFilePath3.set(limitFilePath_3)

                def selectLimitFilePath4():
                    limitFilePath_4 = askopenfilename()
                    limitFilePath4.set(limitFilePath_4)

                def selectLimitFilePath5():
                    limitFilePath_5 = askopenfilename()
                    limitFilePath5.set(limitFilePath_5)

                def selectLimitFilePath6():
                    limitFilePath_6 = askopenfilename()
                    limitFilePath6.set(limitFilePath_6)

                def selectLimitFilePath7():
                    limitFilePath_7 = askopenfilename()
                    limitFilePath7.set(limitFilePath_7)

                def selectLimitFilePath8():
                    limitFilePath_8 = askopenfilename()
                    limitFilePath8.set(limitFilePath_8)

                def selectLimitFilePath9():
                    limitFilePath_9 = askopenfilename()
                    limitFilePath9.set(limitFilePath_9)

            # 宣告 "選擇 Excel" Function
            if True:
                def selectExcelPath():
                    excelPath_ = askopenfilename()
                    excelPath.set(excelPath_)

            # 宣告 "選擇 Test Data 資料夾" Function
            if True:
                def selectFolderPath1():
                    folderPath_1 = askdirectory()
                    folderPath1.set(folderPath_1)

                def selectFolderPath2():
                    folderPath_2 = askdirectory()
                    folderPath2.set(folderPath_2)

                def selectFolderPath3():
                    folderPath_3 = askdirectory()
                    folderPath3.set(folderPath_3)

                def selectFolderPath4():
                    folderPath_4 = askdirectory()
                    folderPath4.set(folderPath_4)

                def selectFolderPath5():
                    folderPath_5 = askdirectory()
                    folderPath5.set(folderPath_5)

                def selectFolderPath6():
                    folderPath_6 = askdirectory()
                    folderPath6.set(folderPath_6)

                def selectFolderPath7():
                    folderPath_7 = askdirectory()
                    folderPath7.set(folderPath_7)

                def selectFolderPath8():
                    folderPath_8 = askdirectory()
                    folderPath8.set(folderPath_8)

                def selectFolderPath9():
                    folderPath_9 = askdirectory()
                    folderPath9.set(folderPath_9)

            # 宣告 "關閉" 按鈕
            def exit_():
                exit()

            # 宣告連線失敗 Delay Retry
            class CommunicationError(Exception):
                pass

            retry_on_communication_error = partial(retry,
                stop = stop_after_delay(20), # max. 20 seconds wait.
                wait = wait_fixed(2), # wait 2 seconds
                retry = retry_if_exception_type(CommunicationError),
                reraise = True)()

        if True: # 建立 GUI

            if True: # 建立選擇 Limit File GUI

                print("請選擇 Limit File 路徑\n")

                # 建立使用者 GUI，使其可選擇 Excel 路徑
                windowWB = tk.Tk()
                windowWB.title("Limit File 選擇")
                windowWB.geometry('600x450')
                windowWB.resizable(0,0)

                limitFilePath1 = StringVar()
                limitFilePath2 = StringVar()
                limitFilePath3 = StringVar()
                limitFilePath4 = StringVar()
                limitFilePath5 = StringVar()
                limitFilePath6 = StringVar()
                limitFilePath7 = StringVar()
                limitFilePath8 = StringVar()
                limitFilePath9 = StringVar()

                # ==================== #

                limitFile_frame1 = tk.Frame(windowWB)
                limitFile_frame1.grid(row=0, column=0, pady=2)

                limitFile_label1 = tk.Label(limitFile_frame1, text="Limit File 路徑1")
                limitFile_label1.grid(row=0, column=0, ipadx=10)
                limitFile_entry1 = tk.Entry(limitFile_frame1, textvariable=limitFilePath1)
                limitFile_entry1.grid(row=0, column=1, ipadx=124)
                limitFile_button1 = tk.Button(limitFile_frame1, text="選擇", command=selectLimitFilePath1)
                limitFile_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame2 = tk.Frame(windowWB)
                limitFile_frame2.grid(row=1, column=0, pady=2)

                limitFile_label2 = tk.Label(limitFile_frame2, text="Limit File 路徑2")
                limitFile_label2.grid(row=0, column=0, ipadx=10)
                limitFile_entry2 = tk.Entry(limitFile_frame2, textvariable=limitFilePath2)
                limitFile_entry2.grid(row=0, column=1, ipadx=124)
                limitFile_button2 = tk.Button(limitFile_frame2, text="選擇", command=selectLimitFilePath2)
                limitFile_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame3 = tk.Frame(windowWB)
                limitFile_frame3.grid(row=2, column=0, pady=2)

                limitFile_label3 = tk.Label(limitFile_frame3, text="Limit File 路徑3")
                limitFile_label3.grid(row=0, column=0, ipadx=10)
                limitFile_entry3 = tk.Entry(limitFile_frame3, textvariable=limitFilePath3)
                limitFile_entry3.grid(row=0, column=1, ipadx=124)
                limitFile_button3 = tk.Button(limitFile_frame3, text="選擇", command=selectLimitFilePath3)
                limitFile_button3.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame4 = tk.Frame(windowWB)
                limitFile_frame4.grid(row=3, column=0, pady=2)

                limitFile_label4 = tk.Label(limitFile_frame4, text="Limit File 路徑4")
                limitFile_label4.grid(row=0, column=0, ipadx=10)
                limitFile_entry4 = tk.Entry(limitFile_frame4, textvariable=limitFilePath4)
                limitFile_entry4.grid(row=0, column=1, ipadx=124)
                limitFile_button4 = tk.Button(limitFile_frame4, text="選擇", command=selectLimitFilePath4)
                limitFile_button4.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame5 = tk.Frame(windowWB)
                limitFile_frame5.grid(row=4, column=0, pady=2)

                limitFile_label5 = tk.Label(limitFile_frame5, text="Limit File 路徑5")
                limitFile_label5.grid(row=0, column=0, ipadx=10)
                limitFile_entry5 = tk.Entry(limitFile_frame5, textvariable=limitFilePath5)
                limitFile_entry5.grid(row=0, column=1, ipadx=124)
                limitFile_button5 = tk.Button(limitFile_frame5, text="選擇", command=selectLimitFilePath5)
                limitFile_button5.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame6 = tk.Frame(windowWB)
                limitFile_frame6.grid(row=5, column=0, pady=2)

                limitFile_label6 = tk.Label(limitFile_frame6, text="Limit File 路徑6")
                limitFile_label6.grid(row=0, column=0, ipadx=10)
                limitFile_entry6 = tk.Entry(limitFile_frame6, textvariable=limitFilePath6)
                limitFile_entry6.grid(row=0, column=1, ipadx=124)
                limitFile_button6 = tk.Button(limitFile_frame6, text="選擇", command=selectLimitFilePath6)
                limitFile_button6.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame7 = tk.Frame(windowWB)
                limitFile_frame7.grid(row=6, column=0, pady=2)

                limitFile_label7 = tk.Label(limitFile_frame7, text="Limit File 路徑7")
                limitFile_label7.grid(row=0, column=0, ipadx=10)
                limitFile_entry7 = tk.Entry(limitFile_frame7, textvariable=limitFilePath7)
                limitFile_entry7.grid(row=0, column=1, ipadx=124)
                limitFile_button7 = tk.Button(limitFile_frame7, text="選擇", command=selectLimitFilePath7)
                limitFile_button7.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame8 = tk.Frame(windowWB)
                limitFile_frame8.grid(row=7, column=0, pady=2)

                limitFile_label8 = tk.Label(limitFile_frame8, text="Limit File 路徑8")
                limitFile_label8.grid(row=0, column=0, ipadx=10)
                limitFile_entry8 = tk.Entry(limitFile_frame8, textvariable=limitFilePath8)
                limitFile_entry8.grid(row=0, column=1, ipadx=124)
                limitFile_button8 = tk.Button(limitFile_frame8, text="選擇", command=selectLimitFilePath8)
                limitFile_button8.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame9 = tk.Frame(windowWB)
                limitFile_frame9.grid(row=8, column=0, pady=2)

                limitFile_label9 = tk.Label(limitFile_frame9, text="Limit File 路徑9")
                limitFile_label9.grid(row=0, column=0, ipadx=10)
                limitFile_entry9 = tk.Entry(limitFile_frame9, textvariable=limitFilePath9)
                limitFile_entry9.grid(row=0, column=1, ipadx=124)
                limitFile_button9 = tk.Button(limitFile_frame9, text="選擇", command=selectLimitFilePath9)
                limitFile_button9.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameWB = tk.Frame(windowWB)
                button_frameWB.grid(row=9, column=0)

                start_buttonWB = tk.Button(button_frameWB, text="確認", width=6, command=windowWB.destroy)
                start_buttonWB.grid(row=0, column=0, padx=8, pady=10)
                exit_buttonWB = tk.Button(button_frameWB, text="關閉", width=6, command=exit_)
                exit_buttonWB.grid(row=0, column=1, padx=8, pady=10)

                windowWB.mainloop()
                limitFilePath = []

                print("Limit File 路徑為：\n")
                if limitFilePath1.get() != "":
                    print(limitFilePath1.get(), "\n")
                    limitFilePath.append(limitFilePath1.get())
                if limitFilePath2.get() != "":
                    print(limitFilePath2.get(), "\n")
                    limitFilePath.append(limitFilePath2.get())
                if limitFilePath3.get() != "":
                    print(limitFilePath3.get(), "\n")
                    limitFilePath.append(limitFilePath3.get())
                if limitFilePath4.get() != "":
                    print(limitFilePath4.get(), "\n")
                    limitFilePath.append(limitFilePath4.get())
                if limitFilePath5.get() != "":
                    print(limitFilePath5.get(), "\n")
                    limitFilePath.append(limitFilePath5.get())
                if limitFilePath6.get() != "":
                    print(limitFilePath6.get(), "\n")
                    limitFilePath.append(limitFilePath6.get())
                if limitFilePath7.get() != "":
                    print(limitFilePath7.get(), "\n")
                    limitFilePath.append(limitFilePath7.get())
                if limitFilePath8.get() != "":
                    print(limitFilePath8.get(), "\n")
                    limitFilePath.append(limitFilePath8.get())
                if limitFilePath9.get() != "":
                    print(limitFilePath9.get(), "\n")
                    limitFilePath.append(limitFilePath9.get())
                print("========================================\n")

            if True: # 建立選擇 Excel GUI

                print("請選擇 Excel 路徑\n")

                # 建立使用者 GUI，使其可選擇 Excel 路徑
                windowWB = tk.Tk()
                windowWB.title("Excel 選擇")
                windowWB.geometry('560x90')
                windowWB.resizable(0,0)

                excelPath  = StringVar()

                excel_frame = tk.Frame(windowWB)
                excel_frame.grid(row=0, column=0, pady=2)

                excel_label = tk.Label(excel_frame, text="Excel 路徑")
                excel_label.grid(row=0, column=0, ipadx=10)
                excel_entry = tk.Entry(excel_frame, textvariable=excelPath)
                excel_entry.grid(row=0, column=1, ipadx=124)
                excel_button = tk.Button(excel_frame, text="選擇", command=selectExcelPath)
                excel_button.grid(row=0, column=2, padx=7, pady=4)

                button_frameWB = tk.Frame(windowWB)
                button_frameWB.grid(row=1, column=0)

                start_buttonWB = tk.Button(button_frameWB, text="確認", width=6, command=windowWB.destroy)
                start_buttonWB.grid(row=0, column=0, padx=8)
                exit_buttonWB = tk.Button(button_frameWB, text="關閉", width=6, command=exit_)
                exit_buttonWB.grid(row=0, column=1, padx=8)

                windowWB.mainloop()

                print("Excel 路徑為：\n")
                print(excelPath.get(), "\n")
                print("========================================\n")

            if True: # 建立導入 Test Data GUI
            
                print("請選擇 Test Data 的路徑\n")

                # 建立使用者 GUI，使其可選擇要導入 Image 的資料夾
                windowImg = tk.Tk()
                windowImg.title("Test Data 路徑選擇")
                windowImg.geometry('600x450')
                windowImg.resizable(0,0)

                folderPath1 = StringVar()
                folderPath2 = StringVar()
                folderPath3 = StringVar()
                folderPath4 = StringVar()
                folderPath5 = StringVar()
                folderPath6 = StringVar()
                folderPath7 = StringVar()
                folderPath8 = StringVar()
                folderPath9 = StringVar()

                # ==================== #

                image_frame1 = tk.Frame(windowImg)
                image_frame1.grid(row=0, column=0, pady=2)

                image_label1 = tk.Label(image_frame1, text="Test Data 路徑1")
                image_label1.grid(row=0, column=0, ipadx=10)
                image_entry1 = tk.Entry(image_frame1, textvariable=folderPath1)
                image_entry1.grid(row=0, column=1, ipadx=124)
                image_button1 = tk.Button(image_frame1, text="選擇", command=selectFolderPath1)
                image_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame2 = tk.Frame(windowImg)
                image_frame2.grid(row=1, column=0, pady=2)

                image_label2 = tk.Label(image_frame2, text="Test Data 路徑2")
                image_label2.grid(row=0, column=0, ipadx=10)
                image_entry2 = tk.Entry(image_frame2, textvariable=folderPath2)
                image_entry2.grid(row=0, column=1, ipadx=124)
                image_button2 = tk.Button(image_frame2, text="選擇", command=selectFolderPath2)
                image_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame3 = tk.Frame(windowImg)
                image_frame3.grid(row=2, column=0, pady=2)

                image_label3 = tk.Label(image_frame3, text="Test Data 路徑3")
                image_label3.grid(row=0, column=0, ipadx=10)
                image_entry3 = tk.Entry(image_frame3, textvariable=folderPath3)
                image_entry3.grid(row=0, column=1, ipadx=124)
                image_button3 = tk.Button(image_frame3, text="選擇", command=selectFolderPath3)
                image_button3.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame4 = tk.Frame(windowImg)
                image_frame4.grid(row=3, column=0, pady=2)

                image_label4 = tk.Label(image_frame4, text="Test Data 路徑4")
                image_label4.grid(row=0, column=0, ipadx=10)
                image_entry4 = tk.Entry(image_frame4, textvariable=folderPath4)
                image_entry4.grid(row=0, column=1, ipadx=124)
                image_button4 = tk.Button(image_frame4, text="選擇", command=selectFolderPath4)
                image_button4.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame5 = tk.Frame(windowImg)
                image_frame5.grid(row=4, column=0, pady=2)

                image_label5 = tk.Label(image_frame5, text="Test Data 路徑5")
                image_label5.grid(row=0, column=0, ipadx=10)
                image_entry5 = tk.Entry(image_frame5, textvariable=folderPath5)
                image_entry5.grid(row=0, column=1, ipadx=124)
                image_button5 = tk.Button(image_frame5, text="選擇", command=selectFolderPath5)
                image_button5.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame6 = tk.Frame(windowImg)
                image_frame6.grid(row=5, column=0, pady=2)

                image_label6 = tk.Label(image_frame6, text="Test Data 路徑6")
                image_label6.grid(row=0, column=0, ipadx=10)
                image_entry6 = tk.Entry(image_frame6, textvariable=folderPath6)
                image_entry6.grid(row=0, column=1, ipadx=124)
                image_button6 = tk.Button(image_frame6, text="選擇", command=selectFolderPath6)
                image_button6.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame7 = tk.Frame(windowImg)
                image_frame7.grid(row=6, column=0, pady=2)

                image_label7 = tk.Label(image_frame7, text="Test Data 路徑7")
                image_label7.grid(row=0, column=0, ipadx=10)
                image_entry7 = tk.Entry(image_frame7, textvariable=folderPath7)
                image_entry7.grid(row=0, column=1, ipadx=124)
                image_button7 = tk.Button(image_frame7, text="選擇", command=selectFolderPath7)
                image_button7.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame8 = tk.Frame(windowImg)
                image_frame8.grid(row=7, column=0, pady=2)

                image_label8 = tk.Label(image_frame8, text="Test Data 路徑8")
                image_label8.grid(row=0, column=0, ipadx=10)
                image_entry8 = tk.Entry(image_frame8, textvariable=folderPath8)
                image_entry8.grid(row=0, column=1, ipadx=124)
                image_button8 = tk.Button(image_frame8, text="選擇", command=selectFolderPath8)
                image_button8.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame9 = tk.Frame(windowImg)
                image_frame9.grid(row=8, column=0, pady=2)

                image_label9 = tk.Label(image_frame9, text="Test Data 路徑9")
                image_label9.grid(row=0, column=0, ipadx=10)
                image_entry9 = tk.Entry(image_frame9, textvariable=folderPath9)
                image_entry9.grid(row=0, column=1, ipadx=124)
                image_button9 = tk.Button(image_frame9, text="選擇", command=selectFolderPath9)
                image_button9.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameImg = tk.Frame(windowImg)
                button_frameImg.grid(row=9, column=0)

                start_buttonImg = tk.Button(button_frameImg, text="確認", width=6, command=windowImg.destroy)
                start_buttonImg.grid(row=0, column=0, padx=8, pady=10)
                exit_buttonImg = tk.Button(button_frameImg, text="關閉", width=6, command=exit_)
                exit_buttonImg.grid(row=0, column=1, padx=8, pady=10)

                windowWB.mainloop()
                folderPath = []

                print("Test Data 路徑為：\n")
                if folderPath1.get() != "":
                    print(folderPath1.get(), "\n")
                    folderPath.append(folderPath1.get())
                if folderPath2.get() != "":
                    print(folderPath2.get(), "\n")
                    folderPath.append(folderPath2.get())
                if folderPath3.get() != "":
                    print(folderPath3.get(), "\n")
                    folderPath.append(folderPath3.get())
                if folderPath4.get() != "":
                    print(folderPath4.get(), "\n")
                    folderPath.append(folderPath4.get())
                if folderPath5.get() != "":
                    print(folderPath5.get(), "\n")
                    folderPath.append(folderPath5.get())
                if folderPath6.get() != "":
                    print(folderPath6.get(), "\n")
                    folderPath.append(folderPath6.get())
                if folderPath7.get() != "":
                    print(folderPath7.get(), "\n")
                    folderPath.append(folderPath7.get())
                if folderPath8.get() != "":
                    print(folderPath8.get(), "\n")
                    folderPath.append(folderPath8.get())
                if folderPath9.get() != "":
                    print(folderPath9.get(), "\n")
                    folderPath.append(folderPath9.get())
                print("========================================\n")

    for executeTime in range(len(limitFilePath)):

        # ================================================================================ #
        #                                                                                  #
        #                          按照 Limit File 編列測項 Excel                           #
        #                                                                                  #
        # ================================================================================ #

        if True:

            # 存取 Excel 內所有工作表的名稱
            ExcelWB = openpyxl.load_workbook(excelPath.get())
            ExcelWS = ExcelWB.sheetnames

            # 存取 Limit File 內所有工作表的名稱
            LimitFileWB = openpyxl.load_workbook(limitFilePath[executeTime])
            LimitFileWS = LimitFileWB.sheetnames

            name = (limitFilePath[executeTime]).split("/")[-1][10:-5]
            if name in ExcelWS:
                editWS = ExcelWB[name]
            else:
                Example = ExcelWB["Example"]
                editWS = ExcelWB.copy_worksheet(Example)
                editWS.title = name

                # 選定要編輯的工作表
                MappingWS = LimitFileWB["SheetNameMapping"]

                normal = ""
                low = ""
                high = ""

                # 定義路徑欄
                for row in MappingWS['A1':'A10']:
                    for cell in row:
                        if cell.value == None:
                            break
                        if cell.value[-4:] == "+25C":
                            normal = MappingWS.cell(row=cell.row, column=2).value
                        if cell.value[-4:] == "-40C":
                            low = MappingWS.cell(row=cell.row, column=2).value
                        if cell.value[-4:] == "+55C":
                            high = MappingWS.cell(row=cell.row, column=2).value

                if normal != "":
                    NormalWS = LimitFileWB[normal]
                    for i in range(1, NormalWS.max_row+1):
                        TS_No = ("TS_#%.3d" % i)
                        Temp = "+25C"
                        Test_Name = NormalWS.cell(row=i, column=7).value
                        Unit = NormalWS.cell(row=i, column=6).value
                        Compare = NormalWS.cell(row=i, column=4).value
                        L_LMT = NormalWS.cell(row=i, column=3).value
                        H_LMT = NormalWS.cell(row=i, column=2).value

                        editWS.cell(row=i+9, column=1).value = TS_No
                        editWS.cell(row=i+9, column=3).value = Temp
                        editWS.cell(row=i+9, column=4).value = Test_Name
                        editWS.cell(row=i+9, column=7).value = Unit
                        editWS.cell(row=i+9, column=8).value = Compare
                        editWS.cell(row=i+9, column=9).value = L_LMT
                        editWS.cell(row=i+9, column=10).value = H_LMT

                        Counter = i

                if low != "":
                    LowWS = LimitFileWB[low]
                    for i in range(1, LowWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+i))
                        Temp = "-40C"
                        Test_Name = LowWS.cell(row=i, column=7).value
                        Unit = LowWS.cell(row=i, column=6).value
                        Compare = LowWS.cell(row=i, column=4).value
                        L_LMT = LowWS.cell(row=i, column=3).value
                        H_LMT = LowWS.cell(row=i, column=2).value

                        editWS.cell(row=Counter+i+9, column=1).value = TS_No
                        editWS.cell(row=Counter+i+9, column=3).value = Temp
                        editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                        editWS.cell(row=Counter+i+9, column=7).value = Unit
                        editWS.cell(row=Counter+i+9, column=8).value = Compare
                        editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                        editWS.cell(row=Counter+i+9, column=10).value = H_LMT

                        LowCount = i

                    Counter = Counter + LowCount
                
                if high != "":
                    HighWS = LimitFileWB[high]
                    for i in range(1, HighWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+i))
                        Temp = "+55C"
                        Test_Name = HighWS.cell(row=i, column=7).value
                        Unit = HighWS.cell(row=i, column=6).value
                        Compare = HighWS.cell(row=i, column=4).value
                        L_LMT = HighWS.cell(row=i, column=3).value
                        H_LMT = HighWS.cell(row=i, column=2).value

                        editWS.cell(row=Counter+i+9, column=1).value = TS_No
                        editWS.cell(row=Counter+i+9, column=3).value = Temp
                        editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                        editWS.cell(row=Counter+i+9, column=7).value = Unit
                        editWS.cell(row=Counter+i+9, column=8).value = Compare
                        editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                        editWS.cell(row=Counter+i+9, column=10).value = H_LMT

                        HighCount = i

                    Counter = Counter + HighCount

        # ================================================================================ #
        #                                                                                  #
        #                                     前置作業                                      #
        #                                                                                  #
        # ================================================================================ #

        if True:
            # 查看 Image 資料夾是否存在，若無，建立 Image 資料夾
            newImagePath = os.path.dirname(excelPath.get())
            if not(os.path.exists(newImagePath + "/Image")):
                os.makedirs(newImagePath + "/Image")

            # 查看測項子 Image 資料夾是否存在，若無，建立測項子 Image 資料夾
            newImageFolder = newImagePath + "/Image/" + name
            if not(os.path.exists(newImageFolder)):
                os.makedirs(newImageFolder)

            # 要編輯的工作表
            # editWS

            # 定義路徑欄
            for row in editWS['A9':'AZ9']:
                for cell in row:
                    unitCoordinate = 2
                    if cell.value == "Temp.":
                        tempCoordinate = cell.column
                    if cell.value == "Test_Name":
                        nameCoordinate = cell.column
                    if cell.value == "Result":
                        resultCoordinate = cell.column
                    if cell.value == "TestVal":
                        valueCoordinate = cell.column
                    if cell.value == "Image":
                        imageCoordinate = cell.column
                    if cell.value == "Session":
                        sessionCoordinate = cell.column
                    if cell.value == "Start_Time":
                        timeCoordinate = cell.column
                    if cell.value == "Exec_Time":
                        execCoordinate = cell.column
                    if cell.value == "A_TestParam":
                        aCoordinate = cell.column
                    if cell.value == "B_TestParam":
                        bCoordinate = cell.column
                    if cell.value == "C_TestParam":
                        cCoordinate = cell.column
                    if cell.value == "D_TestParam":
                        dCoordinate = cell.column
                    if cell.value == "E_TestParam":
                        eCoordinate = cell.column
                    if cell.value == "F_TestParam":
                        fCoordinate = cell.column
                    if cell.value == "G_TestParam":
                        gCoordinate = cell.column
                    if cell.value == "H_TestParam":
                        hCoordinate = cell.column
                    if cell.value == "Validation":
                        validCoordinate = cell.column
                    if cell.value == "Test Plot":
                        picCoordinate = cell.column

            # 宣告最後用來放入 Pic 超連結路徑的 Array
            picPath = ["Path"] * (editWS.max_row + 1 - 10)
            picName = ["Name"] * (editWS.max_row + 1 - 10)
            testPlot = ["Plot"] * (editWS.max_row + 1 - 10)

            # 將所有符合工作表名稱的資料夾放入 folder[]
            folder = []

            for i in range(len(folderPath)):
                dir = ([ name for name in os.listdir(folderPath[i]) if os.path.isdir(os.path.join(folderPath[i], name)) ])
                for j in range(len(dir)):
                    if name == dir[j]:  folder.append(folderPath[i] + "/" + dir[j])

            # 列出所有不同時間、不同溫度之 Test Log
            testLogFolder = []

            for i in range(len(folder)):
                dir = ([ name for name in os.listdir(folder[i]) if os.path.isdir(os.path.join(folder[i], name)) ])
                for j in range(len(dir)):
                    if len(dir[j].split("][")) != 5:
                        continue
                    testLogFolder.append(folder[i] + "/" + dir[j])

            testLogFolder.reverse()
            # 將 Test Log 依時間排序
            folderTimeStr = [""] * len(testLogFolder)
            folderTimeSec = [""] * len(testLogFolder)
            for i in range(len(testLogFolder)):
                folderTimeStr[i] = (testLogFolder[i].split("][")[2]) + "_" + (testLogFolder[i].split("][")[3])
                folderTimeSec[i] = time.mktime((datetime.strptime(folderTimeStr[i], "%Y-%m-%d_%H_%M_%S")).timetuple())

            testLogFolderTime = list(zip(testLogFolder, folderTimeSec))
            sortTestLogFolder = sorted(testLogFolderTime, key= lambda testLogFolderTime : testLogFolderTime[1])
            sortTestLogFolder = np.array(sortTestLogFolder)[:, 0]

        # ================================================================================ #
        #                                                                                  #
        #                                     掃描測項                                      #
        #                                                                                  #
        # ================================================================================ #

        print("掃描檔案，填入 Test Data 與複製照片中...\n")
        print("Running : %s\n" % name)
        startTime = time.time()
        print(time.strftime("Start time : %Y/%m/%d %a %H:%M:%S\n", time.localtime()))

        temp0 = 0
        total0 = editWS.max_row-9
        picDir1 = []
        picDir2 = []
        execTime = time.time()

        @retry_on_communication_error
        def TestDataBuilding(i, count):

            global temp0, picDir1, picDir2, execTime
            temp0 = i-9
            picDir1 = []
            picDir2 = []

            result = []
            for resultCount in range(count):
                # 儲存結果欄位
                result.append(editWS.cell(row=i+resultCount, column=resultCoordinate).value)

            if (None in result) or ("Failed" in result) or ("Skipped" in result):

                testName = []
                newTestName = [""] * count
                compare = 0.5

                for x in range(count):

                    # 儲存名稱欄位
                    testName.append(editWS.cell(row=i+x, column=nameCoordinate).value)

                    # 去掉 Test Name 前面的項目名，只留下後段測試名
                    nameSplit = testName[x].split("][")
                    newTestName[x] = "[" + nameSplit[0][-3:]
                    for y in range(1, len(nameSplit)):
                        newTestName[x] = newTestName[x] + "][" + nameSplit[y]

                # 儲存溫度欄位
                temp = (editWS.cell(row=i, column=tempCoordinate).value)

                for j in range(len(sortTestLogFolder)):
                    # 篩選掉溫度不符合的
                    if temp == ((sortTestLogFolder[j][1:-1]).split("/")[-1]).split("][")[1]:
                        testDir = os.listdir(sortTestLogFolder[j])
                        testDir.sort()
                        for l in range(len(testDir)):
                            if name == testDir[l]:
                                picFolder1 = (sortTestLogFolder[j] + "/" + testDir[l])
                                picDir1 = os.listdir(picFolder1)
                            if testDir[l][-5:] == "_PASS" or testDir[l][-5:] == "_FAIL":
                                if name == testDir[l][:-5]:
                                    picFolder2 = (sortTestLogFolder[j] + "/" + testDir[l])
                                    picDir2 = os.listdir(picFolder2)
                        for k in range(len(testDir)):
                            if testDir[k][-16:] == "_TestReport.xlsx" and testDir[k][:1] == "(" and testDir[k][:2] != "~$":
                                finalFolder = (sortTestLogFolder[j] + "/" + testDir[k])
                                TestDataWB = openpyxl.load_workbook(finalFolder)
                                TestDataWS = TestDataWB[name]
                                for m in range(12, TestDataWS.max_row+1):

                                    # Grouping
                                    trueCounter = 0
                                    for n in range(count):
                                        if editWS.cell(row=i+n, column=8).value == TestDataWS.cell(row=m+n, column=7).value:
                                            if testName[n] == TestDataWS.cell(row=m+n, column=3).value and TestDataWS.cell(row=m+n, column=4).value == "Passed":
                                                trueCounter += 4
                                            if testName[n] == TestDataWS.cell(row=m+n, column=3).value and TestDataWS.cell(row=m+n, column=4).value == "Failed":
                                                trueCounter += 3
                                            if testName[n] == TestDataWS.cell(row=m+n, column=3).value and TestDataWS.cell(row=m+n, column=4).value == "Skipped":
                                                trueCounter += 2
                                            if testName[n] == TestDataWS.cell(row=m+n, column=3).value and TestDataWS.cell(row=m+n, column=4).value == None:
                                                trueCounter += 1

                                    if trueCounter <= count*4 and trueCounter >= compare:
                                        compare = trueCounter

                                        #print(compare)
                                        for o in range(count):
                                            editWS.cell(row=i+o, column=unitCoordinate).value = (testDir[k][1:-1]).split(")(")[0] #dataUnit
                                            editWS.cell(row=i+o, column=resultCoordinate).value = TestDataWS.cell(row=m+o, column=4).value #dataResult
                                            editWS.cell(row=i+o, column=valueCoordinate).value = TestDataWS.cell(row=m+o, column=5).value #dataTestVal

                                            editWS.cell(row=i+o, column=sessionCoordinate).value = finalFolder.split("/")[-2] #dataSession
                                            editWS.cell(row=i+o, column=timeCoordinate).value = TestDataWS.cell(row=m+o, column=11).value #dataStart
                                            editWS.cell(row=i+o, column=execCoordinate).value = TestDataWS.cell(row=m+o, column=12).value #dataExec
                                            editWS.cell(row=i+o, column=aCoordinate).value = TestDataWS.cell(row=m+o, column=13).value #dataA
                                            editWS.cell(row=i+o, column=bCoordinate).value = TestDataWS.cell(row=m+o, column=14).value #dataB
                                            editWS.cell(row=i+o, column=cCoordinate).value = TestDataWS.cell(row=m+o, column=15).value #dataC
                                            editWS.cell(row=i+o, column=dCoordinate).value = TestDataWS.cell(row=m+o, column=16).value #dataD
                                            editWS.cell(row=i+o, column=eCoordinate).value = TestDataWS.cell(row=m+o, column=17).value #dataE
                                            editWS.cell(row=i+o, column=fCoordinate).value = TestDataWS.cell(row=m+o, column=18).value #dataF
                                            editWS.cell(row=i+o, column=gCoordinate).value = TestDataWS.cell(row=m+o, column=19).value #dataG
                                            editWS.cell(row=i+o, column=hCoordinate).value = TestDataWS.cell(row=m+o, column=20).value #dataH
                                            editWS.cell(row=i+o, column=validCoordinate).value = TestDataWS.cell(row=m+o, column=27).value # dataValidation
                                            if TestDataWS.cell(row=m+o, column=29).value != None:
                                                if TestDataWS.cell(row=m+o, column=29).value[-4:] == ".png":
                                                    editWS.cell(row=i+o, column=picCoordinate).value = TestDataWS.cell(row=m+o, column=29).value[:-4] #dataTestPlot
                                                    if (excelPath.get()).split("/")[-1][:2] == "TB" and "RX_NF" in name:
                                                        testPlot[(i+o)-10] = (TestDataWS.cell(row=m+o, column=29).value).split(",")[1]
                                                    else:
                                                        testPlot[(i+o)-10] = TestDataWS.cell(row=m+o, column=29).value
                                                else:
                                                    editWS.cell(row=i+o, column=picCoordinate).value = TestDataWS.cell(row=m+o, column=29).value #dataTestPlot
                                                    if (excelPath.get()).split("/")[-1][:2] == "TB" and "RX_NF" in name:
                                                        testPlot[(i+o)-10] = (TestDataWS.cell(row=m+o, column=29).value).split(",")[1] + ".png"
                                                    else:
                                                        testPlot[(i+o)-10] = TestDataWS.cell(row=m+o, column=29).value + ".png"

                                                # 搜尋符合的照片，記錄其所在資料夾以及名稱
                                                for p in range(len(picDir1)):
                                                    if picDir1[p] == testPlot[(i+o)-10]:
                                                        picPath[(i+o)-10] = picFolder1
                                                        picName[(i+o)-10] = newTestName[o] + "_" + (testDir[k][1:-1]).split(")(")[0][-4:] + ".png"

                                                for r in range(len(picDir2)):
                                                    if picDir2[r] == testPlot[(i+o)-10]:
                                                        picPath[(i+o)-10] = picFolder2
                                                        picName[(i+o)-10] = newTestName[o] + "_" + (testDir[k][1:-1]).split(")(")[0][-4:] + ".png"

                for q in range(count):
                    if picPath[i+q-10] != "Path" and picName[i+q-10] != "Name":
                        # 避免相同的 Test Plot 因 Test Name 不同，而重複複製成兩張不同的照片
                        if testPlot[i+q-10] == testPlot[i+q-11] and (i+q) > 10:
                            picName[i+q-10] = picName[i+q-11]

                        newPath = "Image/" + newImageFolder.split("/")[-1] + "/" + picName[i+q-10]
                        editWS.cell(row=i+q, column=imageCoordinate).value = ('=HYPERLINK("%s", "Link")' % newPath)
                        editWS.cell(row=i+q, column=imageCoordinate).font = Font(underline="single", color='00B050')
                        try:
                            if picName[i+q-10] not in os.listdir(newImageFolder):
                                shutil.copyfile(os.path.join(picPath[i+q-10], testPlot[i+q-10]),os.path.join(newImageFolder, picName[i+q-10]))
                            #pass
                        except:
                            print("         Error.\n")
                            continue

                    if (i+q-10) % 2000 == 0 and i > 10:
                        print("\n\nSaving!", end="")
                        executeTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - execTime))
                        print("     Execute time: %d Hr %d Min %d Sec\n" % (int(executeTime.split("/")[3])-8, int(executeTime.split("/")[4]), int(executeTime.split("/")[5])))
                        execTime = time.time()
                        ExcelWB.save(excelPath.get())

                    # Progress Bar ( 掃描檔案，填入 Test Data )
                    print('\r' + '%.2f%% |%s%s| %d/%d' % ((float((temp0+q)/total0*100)), '█' * int((temp0+q)*50/total0), ' ' * (50-int((temp0+q)*50/total0)), (i+q-9), (editWS.max_row-9)), end='')

        # Grouping Count
        if True:

            if "Power_Consump" in name:
                groupCount = 1

            # TX Grouping Count
            if "TX_OutputPower_SCA" in name:
                groupCount = 5
            if "TX_OutputPower_DCA" in name:
                groupCount = 4
            if "TX_TPDR" in name:
                groupCount = 1
            if "TX_MODQual" in name:
                groupCount = 3
            if "TX_EVM" in name:
                groupCount = 3
            if "TX_TAE" in name:
                groupCount = 1
            if "TX_OBW" in name:
                groupCount = 1
            if "TX_ACLR" in name:
                groupCount = 5
            if "TX_OBUE" in name:
                groupCount = 7
            if "TX_Spur" in name:
                groupCount = 2

            # TX_IM Grouping Count
            if "TX_IM_ACLR" in name:
                groupCount = 5
            if "TX_IM_OBUE" in name:
                groupCount = 7
            if "TX_IM_Spur" in name:
                groupCount = 2

            # RX Grouping Count
            if "RX_RSSI" in name:
                groupCount = 1
            if "RX_GA" in name:
                groupCount = 1
            if "RX_LSNR" in name:
                groupCount = 1
            if "RX_NF" in name:
                groupCount = 1
            if "RX_ACS" in name:
                groupCount = 1
            if "RX_IBB" in name:
                groupCount = 1
            if "RX_NBB" in name:
                groupCount = 1
            if "RX_OOB" in name:
                groupCount = 1
            if "RX_COL" in name:
                groupCount = 1
            if "RX_IM" in name:
                groupCount = 1

        for i in range(10, editWS.max_row+1, groupCount):
            while True:
                try:
                    TestDataBuilding(i, groupCount)
                    break
                except:
                    print("\n\nConnection Error.    Retry.\n")
                    time.sleep(2)
                    #continue

        # ================================================================================ #
        #                                                                                  #
        #                                檔案儲存 ( Excel )                                 #
        #                                                                                  #
        # ================================================================================ #

        if True:

            resultRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",E3)))']
            editWS.conditional_formatting.add('E3', resultRulePassed)
            resultRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",E3)))']
            editWS.conditional_formatting.add('E3', resultRuleFailed)

            dataCell = 'E10:E%d' % (editWS.max_row)
            dataRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            dataRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",E10)))']
            editWS.conditional_formatting.add(dataCell, dataRulePassed)
            dataRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            dataRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",E10)))']
            editWS.conditional_formatting.add(dataCell, dataRuleFailed)

            editWS.cell(row=3, column=2).value = name
            editWS.cell(row=3, column=2).font = Font(name='Calibri', size=9)
            editWS.cell(row=4, column=5).value = ('=SUBTOTAL(103,$E$10:$E$%d)' % editWS.max_row)
            editWS.cell(row=6, column=5).value = ('=SUMPRODUCT(SUBTOTAL(3,OFFSET($E$10:$E$%d,ROW($E$10:$E$%d)-MIN(ROW($E$10:$E$%d)),,1))*($E$10:$E$%d="Passed"))' % (editWS.max_row, editWS.max_row, editWS.max_row, editWS.max_row))

            resultWS = ExcelWB["Results_Overview"]
            coverageWS = ExcelWB["Coverage_Overview"]
            resultWS.cell(row=resultWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            resultWS.cell(row=resultWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            resultWS.cell(row=resultWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            coverageWS.cell(row=coverageWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            coverageWS.cell(row=coverageWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))

            print("\n\nSaving...  Please Wait...\n")
            print("========================================\n")
            #wb.save(excelPath.get()[:-5] + "_copy.xlsx")
            while True:
                try:
                    ExcelWB.save(excelPath.get())
                    break
                except:
                    print("\n\nConnection Error.    Retry Saving.\n")
                    time.sleep(2)
            print("Success!\n")
            blockTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - startTime))
            totalTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - totalStartTime))
            print("Block execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(blockTime.split("/")[2])-1, int(blockTime.split("/")[3])-8, int(blockTime.split("/")[4]), int(blockTime.split("/")[5])))
            print("Total execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(totalTime.split("/")[2])-1, int(totalTime.split("/")[3])-8, int(totalTime.split("/")[4]), int(totalTime.split("/")[5])))
            print("========================================\n")
            time.sleep(2)

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    if True:

        print("Re-run Y/N ?")
        executeYN = input()
        print()
        print("========================================\n")

        if executeYN == "Y" or executeYN == "y":
            execute = True
        elif executeYN == "N" or executeYN == "n":
            execute = False
        if execute == False:
            break
