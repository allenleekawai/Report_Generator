# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import getch
import pathlib

import openpyxl
from openpyxl.styles import Font

import time

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

execute = True

while True:

    # 宣告 "關閉" 按鈕
    def exit_():
        sys.exit()

    def selectExcelPath():
        excelPath_ = askopenfilename()
        excelPath.set(excelPath_)

    # ================================================================================ #
    #                                                                                  #
    #                             宣告 Function 及建立 GUI                              #
    #                                                                                  #
    # ================================================================================ #

    if True:

        if True: # 宣告 Function

            # 宣告 "選擇 Excel" Function
            if True:
                def selectExcelPathNew():
                    excelPathNew_ = askdirectory()
                    excelPathNew.set(excelPathNew_)

                def selectExcelPathOld():
                    excelPathOld_ = askopenfilename()
                    excelPathOld.set(excelPathOld_)

            # 宣告 "選擇 Config" Function
            if True:
                def selectConfigPath(event):
                    configPath.set('' + configPathText.get())

            # 宣告 "選擇 Phase" Function
            if True:
                def selectPhase(event):
                    phaseSelect.set('' + phaseSelectText.get())

            # 宣告 "關閉" 按鈕
            def exit_():
                sys.exit()

            # 宣告 "cmd暫停"
            def pause():
                print("\nPress any key to continue . . . \n")
                getch.getch()
                exit_()

            # 宣告連線失敗 Delay Retry
            class CommunicationError(Exception):
                pass

        if True: # 建立 GUI

            if True: # 建立選擇 Excel GUI

                print("Please select \"Excel file\" path\n")

                # 建立使用者 GUI，使其可選擇 Excel 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Excel path")
                windowWB.geometry('600x120')
                windowWB.resizable(0,0)

                excelPathNew = StringVar()
                excelPathOld = StringVar()

                # ==================== #

                excel_frame1 = tk.Frame(windowWB)
                excel_frame1.grid(row=0, column=0, pady=2)

                excel_label1 = tk.Label(excel_frame1, text="New Excel path")
                excel_label1.grid(row=0, column=0, ipadx=10)
                excel_entry1 = tk.Entry(excel_frame1, textvariable=excelPathNew)
                excel_entry1.grid(row=0, column=1, ipadx=124)
                excel_button1 = tk.Button(excel_frame1, text="Select", width=6, command=selectExcelPathNew)
                excel_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                excel_frame2 = tk.Frame(windowWB)
                excel_frame2.grid(row=1, column=0, pady=2)

                excel_label2 = tk.Label(excel_frame2, text=" Old Excel path")
                excel_label2.grid(row=0, column=0, ipadx=10)
                excel_entry2 = tk.Entry(excel_frame2, textvariable=excelPathOld)
                excel_entry2.grid(row=0, column=1, ipadx=124)
                excel_button2 = tk.Button(excel_frame2, text="Select", width=6, command=selectExcelPathOld)
                excel_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameEX = tk.Frame(windowWB)
                button_frameEX.grid(row=2, column=0)

                start_buttonEX = tk.Button(button_frameEX, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonEX.grid(row=0, column=0, padx=10)
                exit_buttonEX = tk.Button(button_frameEX, text="Close", width=9, command=exit_)
                exit_buttonEX.grid(row=0, column=1, padx=10)

                windowWB.mainloop()
                excelPath = ""

                print("Excel path is : \n")
                if excelPathNew.get() != "":
                    print(excelPathNew.get(), "\n")
                    excelPath = excelPathNew.get()
                if excelPathOld.get() != "":
                    print(excelPathOld.get(), "\n")
                    excelPath = excelPathOld.get()
                print("========================================\n")

            if True: # 建立選擇 Config GUI

                print("Please select \"Config\"\n")

                # 建立使用者 GUI，使其可選擇 Config 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Config")
                windowWB.geometry('370x140')
                windowWB.resizable(0,0)

                configPathText = tk.StringVar()

                config_frame = tk.Frame(windowWB)
                config_frame.grid(row=0, column=0, pady=2)

                config_label = tk.Label(config_frame, text="Select Config")
                config_label.grid(row=0, column=0, ipadx=10, ipady=3, pady=5)
                config_combo = ttk.Combobox(config_frame, textvariable=configPathText, state="readonly")
                config_combo['values'] = ["DB G1.0", "DB G1.5", "TB G1.0", "TB G1.5", "Quadriga B1", "Quadriga B7", "NEC 4GLTE", "NEC 5GNR"]
                config_combo.grid(row=1, column=0, ipadx=70, ipady=1, padx=10)
                #config_combo.current(1)

                button_frameCF = tk.Frame(windowWB)
                button_frameCF.grid(row=1, column=0, pady=15)

                config_combo.bind('<<ComboboxSelected>>', selectConfigPath)
                configPath = tk.StringVar()

                start_buttonCF = tk.Button(button_frameCF, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonCF.grid(row=0, column=0, padx=10)
                exit_buttonCF = tk.Button(button_frameCF, text="Close", width=9, command=exit_)
                exit_buttonCF.grid(row=0, column=1, padx=10)

                windowWB.mainloop()

                print("Config path is : \n")
                print(configPath.get(), "\n")
                configPath = configPath.get()
                print("========================================\n")

            if True: # 建立選擇 Phase GUI

                print("Please select \"Phase\"\n")

                # 建立使用者 GUI，使其可選擇 Phase 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Phase")
                windowWB.geometry('370x140')
                windowWB.resizable(0,0)

                phaseSelectText = tk.StringVar()

                phase_frame = tk.Frame(windowWB)
                phase_frame.grid(row=0, column=0, pady=2)

                phase_label = tk.Label(phase_frame, text="Phase")
                phase_label.grid(row=0, column=0, ipadx=10, ipady=3, pady=5)
                phase_combo = ttk.Combobox(phase_frame, textvariable=phaseSelectText, state="readonly")
                phase_combo['values'] = ["Phase 0", "Phase 1", "Phase 2"]
                phase_combo.grid(row=1, column=0, ipadx=70, ipady=1, padx=10)
                #phase_combo.current(1)

                button_framePH = tk.Frame(windowWB)
                button_framePH.grid(row=1, column=0, pady=15)

                phase_combo.bind('<<ComboboxSelected>>', selectPhase)
                phaseSelect = tk.StringVar()

                start_buttonPH = tk.Button(button_framePH, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonPH.grid(row=0, column=0, padx=10)
                exit_buttonPH = tk.Button(button_framePH, text="Close", width=9, command=exit_)
                exit_buttonPH.grid(row=0, column=1, padx=10)

                windowWB.mainloop()

                print("Phase select is : \n")
                print(phaseSelect.get(), "\n")
                if phaseSelect.get() == "Phase 0":
                    phaseSelect = "P0"
                elif phaseSelect.get() == "Phase 1":
                    phaseSelect = "P1"
                elif phaseSelect.get() == "Phase 2":
                    phaseSelect = "P2"
                print("========================================\n")

    # 打開 Config
    configPathTxt = configPath + " Config.txt"
    txt = open(configPathTxt, "r")
    configArray = (txt.read()).split("\n")
    virtual = configArray.pop(0)
    for configList in range(len(configArray)):
        if "" in configArray:
            configArray.remove("")
    
    print("Updating Server Hyperlink ...  Please Wait...\n")
    print("========================================\n")

    rightNowPath = os.path.join(pathlib.Path().absolute())

    ExcelWB = openpyxl.load_workbook(excelPath)

    ExcelWS = ExcelWB.sheetnames
    CheckWS = ExcelWS

    
    if "Results_Overview" in CheckWS:
        CheckWS.remove("Results_Overview")
    if "Coverage_Overview" in CheckWS:
        CheckWS.remove("Coverage_Overview")
    if "Example" in CheckWS:
        CheckWS.remove("Example")
    if "SpcExample" in CheckWS:
        CheckWS.remove("SpcExample")

    for name in CheckWS:

        editWS = ExcelWB[name]

        # ================================================================================ #
        #                                                                                  #
        #                                檔案儲存 ( Excel )                                 #
        #                                                                                  #
        # ================================================================================ #

        if True:

            for row in editWS['A9':'AZ9']:
                for cell in row:
                    if cell.value == "Image":
                        imageLocalCoordinate = cell.column
                    #if cell.value == "Image_Server":
                    #    imageServerCoordinate = cell.column

            editWS.insert_cols(12)
            editWS['K9'].value = "Image_Local"
            editWS['L9'].value = "Image_Server"

            for rowCounter in range(10, editWS.max_row):
                LocalHyper = editWS.cell(row=rowCounter, column=imageLocalCoordinate).value
                if LocalHyper != None:
                    print(LocalHyper)
                    LocalHyperCut = LocalHyper[18:-10]
                    ServerHyper = '=HYPERLINK("https://twpub10.mtigroup.com/' + virtual + '/Image/' + configPath + '/' + phaseSelect + '/' + LocalHyperCut + '", "Remote_Link")'
                    editWS.cell(row=rowCounter, column=imageLocalCoordinate+1).value = ServerHyper
                    editWS.cell(row=rowCounter, column=imageLocalCoordinate+1).font = Font(underline="single", color='00B050')

    print("Saving...            Please Wait...\n")
    #wb.save(excelPath.get()[:-5] + "_copy.xlsx")
    saveErrorCount = 0
    while True:
        try:
            ExcelWB.save(excelPath)
            break
        except:
            print("Connection Error.    Retry Saving.\n")
            saveErrorCount += 1
            if saveErrorCount == 5:
                print("Timeout, Program stop, Saving is not finish")
                time.sleep(5)
                exit_()
            time.sleep(2)
    print("Success!\n")
    print("========================================\n")
    time.sleep(2)

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    if True:

        print("Re-run Y/N ?")
        executeYN = input()
        print()
        print("========================================\n")

        if executeYN == "Y" or executeYN == "y":
            execute = True
        elif executeYN == "N" or executeYN == "n":
            execute = False
        if execute == False:
            break
