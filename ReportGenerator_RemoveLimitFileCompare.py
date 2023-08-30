# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import getch
import shutil
import pathlib
import numpy as np

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle

import time
from datetime import datetime

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

from functools import partial
from tenacity import retry, stop_after_delay, wait_fixed, retry_if_exception_type

execute = True

while True:

    totalStartTime = time.time() # 紀錄開始時間

    # ================================================================================ #
    #                                                                                  #
    #                             宣告 Function 及建立 GUI                              #
    #                                                                                  #
    # ================================================================================ #

    if True:

        if True: # 宣告 Function

            # 宣告 "選擇 Limit File" Function
            if True:
                def selectLimitFilePath1():
                    limitFilePath_1 = askopenfilename()
                    limitFilePath1.set(limitFilePath_1)

                def selectLimitFilePath2():
                    limitFilePath_2 = askopenfilename()
                    limitFilePath2.set(limitFilePath_2)

                def selectLimitFilePath3():
                    limitFilePath_3 = askopenfilename()
                    limitFilePath3.set(limitFilePath_3)

                def selectLimitFilePath4():
                    limitFilePath_4 = askopenfilename()
                    limitFilePath4.set(limitFilePath_4)

                def selectLimitFilePath5():
                    limitFilePath_5 = askopenfilename()
                    limitFilePath5.set(limitFilePath_5)

                def selectLimitFilePath6():
                    limitFilePath_6 = askopenfilename()
                    limitFilePath6.set(limitFilePath_6)

                def selectLimitFilePath7():
                    limitFilePath_7 = askopenfilename()
                    limitFilePath7.set(limitFilePath_7)

                def selectLimitFilePath8():
                    limitFilePath_8 = askopenfilename()
                    limitFilePath8.set(limitFilePath_8)

                def selectLimitFilePath9():
                    limitFilePath_9 = askopenfilename()
                    limitFilePath9.set(limitFilePath_9)

            # 宣告 "選擇 Excel" Function
            if True:
                def selectExcelPathNew():
                    excelPathNew_ = askdirectory()
                    excelPathNew.set(excelPathNew_)

                def selectExcelPathOld():
                    excelPathOld_ = askopenfilename()
                    excelPathOld.set(excelPathOld_)

            # 宣告 "選擇 Config" Function
            if True:
                def selectConfigPath(event):
                    configPath.set('' + configPathText.get())

            # 宣告 "選擇 Phase" Function
            if True:
                def selectPhase(event):
                    phaseSelect.set('' + phaseSelectText.get())

            # 宣告 "選擇 Test Data 資料夾" Function
            if True:
                def selectFolderPath1():
                    folderPath_1 = askdirectory()
                    folderPath1.set(folderPath_1)

                def selectFolderPath2():
                    folderPath_2 = askdirectory()
                    folderPath2.set(folderPath_2)

                def selectFolderPath3():
                    folderPath_3 = askdirectory()
                    folderPath3.set(folderPath_3)

                def selectFolderPath4():
                    folderPath_4 = askdirectory()
                    folderPath4.set(folderPath_4)

                def selectFolderPath5():
                    folderPath_5 = askdirectory()
                    folderPath5.set(folderPath_5)

                def selectFolderPath6():
                    folderPath_6 = askdirectory()
                    folderPath6.set(folderPath_6)

                def selectFolderPath7():
                    folderPath_7 = askdirectory()
                    folderPath7.set(folderPath_7)

                def selectFolderPath8():
                    folderPath_8 = askdirectory()
                    folderPath8.set(folderPath_8)

                def selectFolderPath9():
                    folderPath_9 = askdirectory()
                    folderPath9.set(folderPath_9)

            # 宣告 "關閉" 按鈕
            def exit_():
                sys.exit()

            # 宣告 "cmd暫停"
            def pause():
                print("\nPress any key to continue . . . \n")
                getch.getch()
                exit_()

            # 宣告連線失敗 Delay Retry
            class CommunicationError(Exception):
                pass

            retry_on_communication_error = partial(retry,
                stop = stop_after_delay(20), # max. 20 seconds wait.
                wait = wait_fixed(2), # wait 2 seconds
                retry = retry_if_exception_type(CommunicationError),
                reraise = True)()

        if True: # 建立 GUI

            if True: # 建立選擇 Limit File GUI

                print("Please select \"Limit File\" path\n")

                # 建立使用者 GUI，使其可選擇 Limit File 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Limit File")
                windowWB.geometry('600x450')
                windowWB.resizable(0,0)

                limitFilePath1 = StringVar()
                limitFilePath2 = StringVar()
                limitFilePath3 = StringVar()
                limitFilePath4 = StringVar()
                limitFilePath5 = StringVar()
                limitFilePath6 = StringVar()
                limitFilePath7 = StringVar()
                limitFilePath8 = StringVar()
                limitFilePath9 = StringVar()

                # ==================== #

                limitFile_frame1 = tk.Frame(windowWB)
                limitFile_frame1.grid(row=0, column=0, pady=2)

                limitFile_label1 = tk.Label(limitFile_frame1, text="Limit File Path 1")
                limitFile_label1.grid(row=0, column=0, ipadx=10)
                limitFile_entry1 = tk.Entry(limitFile_frame1, textvariable=limitFilePath1)
                limitFile_entry1.grid(row=0, column=1, ipadx=124)
                limitFile_button1 = tk.Button(limitFile_frame1, text="Select", width=6, command=selectLimitFilePath1)
                limitFile_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame2 = tk.Frame(windowWB)
                limitFile_frame2.grid(row=1, column=0, pady=2)

                limitFile_label2 = tk.Label(limitFile_frame2, text="Limit File Path 2")
                limitFile_label2.grid(row=0, column=0, ipadx=10)
                limitFile_entry2 = tk.Entry(limitFile_frame2, textvariable=limitFilePath2)
                limitFile_entry2.grid(row=0, column=1, ipadx=124)
                limitFile_button2 = tk.Button(limitFile_frame2, text="Select", width=6, command=selectLimitFilePath2)
                limitFile_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame3 = tk.Frame(windowWB)
                limitFile_frame3.grid(row=2, column=0, pady=2)

                limitFile_label3 = tk.Label(limitFile_frame3, text="Limit File Path 3")
                limitFile_label3.grid(row=0, column=0, ipadx=10)
                limitFile_entry3 = tk.Entry(limitFile_frame3, textvariable=limitFilePath3)
                limitFile_entry3.grid(row=0, column=1, ipadx=124)
                limitFile_button3 = tk.Button(limitFile_frame3, text="Select", width=6, command=selectLimitFilePath3)
                limitFile_button3.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame4 = tk.Frame(windowWB)
                limitFile_frame4.grid(row=3, column=0, pady=2)

                limitFile_label4 = tk.Label(limitFile_frame4, text="Limit File Path 4")
                limitFile_label4.grid(row=0, column=0, ipadx=10)
                limitFile_entry4 = tk.Entry(limitFile_frame4, textvariable=limitFilePath4)
                limitFile_entry4.grid(row=0, column=1, ipadx=124)
                limitFile_button4 = tk.Button(limitFile_frame4, text="Select", width=6, command=selectLimitFilePath4)
                limitFile_button4.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame5 = tk.Frame(windowWB)
                limitFile_frame5.grid(row=4, column=0, pady=2)

                limitFile_label5 = tk.Label(limitFile_frame5, text="Limit File Path 5")
                limitFile_label5.grid(row=0, column=0, ipadx=10)
                limitFile_entry5 = tk.Entry(limitFile_frame5, textvariable=limitFilePath5)
                limitFile_entry5.grid(row=0, column=1, ipadx=124)
                limitFile_button5 = tk.Button(limitFile_frame5, text="Select", width=6, command=selectLimitFilePath5)
                limitFile_button5.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame6 = tk.Frame(windowWB)
                limitFile_frame6.grid(row=5, column=0, pady=2)

                limitFile_label6 = tk.Label(limitFile_frame6, text="Limit File Path 6")
                limitFile_label6.grid(row=0, column=0, ipadx=10)
                limitFile_entry6 = tk.Entry(limitFile_frame6, textvariable=limitFilePath6)
                limitFile_entry6.grid(row=0, column=1, ipadx=124)
                limitFile_button6 = tk.Button(limitFile_frame6, text="Select", width=6, command=selectLimitFilePath6)
                limitFile_button6.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame7 = tk.Frame(windowWB)
                limitFile_frame7.grid(row=6, column=0, pady=2)

                limitFile_label7 = tk.Label(limitFile_frame7, text="Limit File Path 7")
                limitFile_label7.grid(row=0, column=0, ipadx=10)
                limitFile_entry7 = tk.Entry(limitFile_frame7, textvariable=limitFilePath7)
                limitFile_entry7.grid(row=0, column=1, ipadx=124)
                limitFile_button7 = tk.Button(limitFile_frame7, text="Select", width=6, command=selectLimitFilePath7)
                limitFile_button7.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame8 = tk.Frame(windowWB)
                limitFile_frame8.grid(row=7, column=0, pady=2)

                limitFile_label8 = tk.Label(limitFile_frame8, text="Limit File Path 8")
                limitFile_label8.grid(row=0, column=0, ipadx=10)
                limitFile_entry8 = tk.Entry(limitFile_frame8, textvariable=limitFilePath8)
                limitFile_entry8.grid(row=0, column=1, ipadx=124)
                limitFile_button8 = tk.Button(limitFile_frame8, text="Select", width=6, command=selectLimitFilePath8)
                limitFile_button8.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                limitFile_frame9 = tk.Frame(windowWB)
                limitFile_frame9.grid(row=8, column=0, pady=2)

                limitFile_label9 = tk.Label(limitFile_frame9, text="Limit File Path 9")
                limitFile_label9.grid(row=0, column=0, ipadx=10)
                limitFile_entry9 = tk.Entry(limitFile_frame9, textvariable=limitFilePath9)
                limitFile_entry9.grid(row=0, column=1, ipadx=124)
                limitFile_button9 = tk.Button(limitFile_frame9, text="Select", width=6, command=selectLimitFilePath9)
                limitFile_button9.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameLF = tk.Frame(windowWB)
                button_frameLF.grid(row=9, column=0)

                start_buttonLF = tk.Button(button_frameLF, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonLF.grid(row=0, column=0, padx=10, pady=10)
                exit_buttonLF = tk.Button(button_frameLF, text="Close", width=9, command=exit_)
                exit_buttonLF.grid(row=0, column=1, padx=10, pady=10)

                windowWB.mainloop()
                limitFilePath = []

                print("Limit File path is : \n")
                if limitFilePath1.get() != "":
                    print(limitFilePath1.get(), "\n")
                    limitFilePath.append(limitFilePath1.get())
                if limitFilePath2.get() != "":
                    print(limitFilePath2.get(), "\n")
                    limitFilePath.append(limitFilePath2.get())
                if limitFilePath3.get() != "":
                    print(limitFilePath3.get(), "\n")
                    limitFilePath.append(limitFilePath3.get())
                if limitFilePath4.get() != "":
                    print(limitFilePath4.get(), "\n")
                    limitFilePath.append(limitFilePath4.get())
                if limitFilePath5.get() != "":
                    print(limitFilePath5.get(), "\n")
                    limitFilePath.append(limitFilePath5.get())
                if limitFilePath6.get() != "":
                    print(limitFilePath6.get(), "\n")
                    limitFilePath.append(limitFilePath6.get())
                if limitFilePath7.get() != "":
                    print(limitFilePath7.get(), "\n")
                    limitFilePath.append(limitFilePath7.get())
                if limitFilePath8.get() != "":
                    print(limitFilePath8.get(), "\n")
                    limitFilePath.append(limitFilePath8.get())
                if limitFilePath9.get() != "":
                    print(limitFilePath9.get(), "\n")
                    limitFilePath.append(limitFilePath9.get())
                print("========================================\n")

            if True: # 建立選擇 Excel GUI

                print("Please select \"Excel file\" path\n")

                # 建立使用者 GUI，使其可選擇 Excel 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Excel path")
                windowWB.geometry('600x120')
                windowWB.resizable(0,0)

                excelPathNew = StringVar()
                excelPathOld = StringVar()

                # ==================== #

                excel_frame1 = tk.Frame(windowWB)
                excel_frame1.grid(row=0, column=0, pady=2)

                excel_label1 = tk.Label(excel_frame1, text="New Excel path")
                excel_label1.grid(row=0, column=0, ipadx=10)
                excel_entry1 = tk.Entry(excel_frame1, textvariable=excelPathNew)
                excel_entry1.grid(row=0, column=1, ipadx=124)
                excel_button1 = tk.Button(excel_frame1, text="Select", width=6, command=selectExcelPathNew)
                excel_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                excel_frame2 = tk.Frame(windowWB)
                excel_frame2.grid(row=1, column=0, pady=2)

                excel_label2 = tk.Label(excel_frame2, text=" Old Excel path")
                excel_label2.grid(row=0, column=0, ipadx=10)
                excel_entry2 = tk.Entry(excel_frame2, textvariable=excelPathOld)
                excel_entry2.grid(row=0, column=1, ipadx=124)
                excel_button2 = tk.Button(excel_frame2, text="Select", width=6, command=selectExcelPathOld)
                excel_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameEX = tk.Frame(windowWB)
                button_frameEX.grid(row=2, column=0)

                start_buttonEX = tk.Button(button_frameEX, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonEX.grid(row=0, column=0, padx=10)
                exit_buttonEX = tk.Button(button_frameEX, text="Close", width=9, command=exit_)
                exit_buttonEX.grid(row=0, column=1, padx=10)

                windowWB.mainloop()
                excelPath = ""

                print("Excel path is : \n")
                if excelPathNew.get() != "":
                    print(excelPathNew.get(), "\n")
                    excelPath = excelPathNew.get()
                if excelPathOld.get() != "":
                    print(excelPathOld.get(), "\n")
                    excelPath = excelPathOld.get()
                print("========================================\n")

            if True: # 建立選擇 Config GUI

                print("Please select \"Config\"\n")

                # 建立使用者 GUI，使其可選擇 Config 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Config")
                windowWB.geometry('370x140')
                windowWB.resizable(0,0)

                configPathText = tk.StringVar()

                config_frame = tk.Frame(windowWB)
                config_frame.grid(row=0, column=0, pady=2)

                config_label = tk.Label(config_frame, text="Select Config")
                config_label.grid(row=0, column=0, ipadx=10, ipady=3, pady=5)
                config_combo = ttk.Combobox(config_frame, textvariable=configPathText, state="readonly")
                config_combo['values'] = ["DB G1.0", "DB G1.5", "TB G1.0", "TB G1.5", "TB G1.5_Fingu", "Quadriga B1", "Quadriga B7", "NEC 4GLTE", "NEC 5GNR"]
                config_combo.grid(row=1, column=0, ipadx=70, ipady=1, padx=10)
                #config_combo.current(1)

                button_frameCF = tk.Frame(windowWB)
                button_frameCF.grid(row=1, column=0, pady=15)

                config_combo.bind('<<ComboboxSelected>>', selectConfigPath)
                configPath = tk.StringVar()

                start_buttonCF = tk.Button(button_frameCF, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonCF.grid(row=0, column=0, padx=10)
                exit_buttonCF = tk.Button(button_frameCF, text="Close", width=9, command=exit_)
                exit_buttonCF.grid(row=0, column=1, padx=10)

                windowWB.mainloop()

                print("Config path is : \n")
                print(configPath.get(), "\n")
                configPath = configPath.get()
                print("========================================\n")

            if True: # 建立選擇 Phase GUI

                print("Please select \"Phase\"\n")

                # 建立使用者 GUI，使其可選擇 Phase 路徑
                windowWB = tk.Tk()
                windowWB.title("Select Phase")
                windowWB.geometry('370x140')
                windowWB.resizable(0,0)

                phaseSelectText = tk.StringVar()

                phase_frame = tk.Frame(windowWB)
                phase_frame.grid(row=0, column=0, pady=2)

                phase_label = tk.Label(phase_frame, text="Phase")
                phase_label.grid(row=0, column=0, ipadx=10, ipady=3, pady=5)
                phase_combo = ttk.Combobox(phase_frame, textvariable=phaseSelectText, state="readonly")
                phase_combo['values'] = ["Phase 0", "Phase 1", "Phase 2"]
                phase_combo.grid(row=1, column=0, ipadx=70, ipady=1, padx=10)
                #phase_combo.current(1)

                button_framePH = tk.Frame(windowWB)
                button_framePH.grid(row=1, column=0, pady=15)

                phase_combo.bind('<<ComboboxSelected>>', selectPhase)
                phaseSelect = tk.StringVar()

                start_buttonPH = tk.Button(button_framePH, text="Confirm", width=9, command=windowWB.destroy)
                start_buttonPH.grid(row=0, column=0, padx=10)
                exit_buttonPH = tk.Button(button_framePH, text="Close", width=9, command=exit_)
                exit_buttonPH.grid(row=0, column=1, padx=10)

                windowWB.mainloop()

                print("Phase select is : \n")
                print(phaseSelect.get(), "\n")
                if phaseSelect.get() == "Phase 0":
                    phaseSelect = "P0"
                elif phaseSelect.get() == "Phase 1":
                    phaseSelect = "P1"
                elif phaseSelect.get() == "Phase 2":
                    phaseSelect = "P2"
                print("========================================\n")

            if True: # 建立導入 Test Data GUI

                print("Please select \"Test Data\" path\n")

                # 建立使用者 GUI，使其可選擇要導入 Test Data 的資料夾
                windowImg = tk.Tk()
                windowImg.title("Select Test Data path")
                windowImg.geometry('600x450')
                windowImg.resizable(0,0)

                folderPath1 = StringVar()
                folderPath2 = StringVar()
                folderPath3 = StringVar()
                folderPath4 = StringVar()
                folderPath5 = StringVar()
                folderPath6 = StringVar()
                folderPath7 = StringVar()
                folderPath8 = StringVar()
                folderPath9 = StringVar()

                # ==================== #

                image_frame1 = tk.Frame(windowImg)
                image_frame1.grid(row=0, column=0, pady=2)

                image_label1 = tk.Label(image_frame1, text="Test Data path 1")
                image_label1.grid(row=0, column=0, ipadx=10)
                image_entry1 = tk.Entry(image_frame1, textvariable=folderPath1)
                image_entry1.grid(row=0, column=1, ipadx=124)
                image_button1 = tk.Button(image_frame1, text="Select", width=6, command=selectFolderPath1)
                image_button1.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame2 = tk.Frame(windowImg)
                image_frame2.grid(row=1, column=0, pady=2)

                image_label2 = tk.Label(image_frame2, text="Test Data path 2")
                image_label2.grid(row=0, column=0, ipadx=10)
                image_entry2 = tk.Entry(image_frame2, textvariable=folderPath2)
                image_entry2.grid(row=0, column=1, ipadx=124)
                image_button2 = tk.Button(image_frame2, text="Select", width=6, command=selectFolderPath2)
                image_button2.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame3 = tk.Frame(windowImg)
                image_frame3.grid(row=2, column=0, pady=2)

                image_label3 = tk.Label(image_frame3, text="Test Data path 3")
                image_label3.grid(row=0, column=0, ipadx=10)
                image_entry3 = tk.Entry(image_frame3, textvariable=folderPath3)
                image_entry3.grid(row=0, column=1, ipadx=124)
                image_button3 = tk.Button(image_frame3, text="Select", width=6, command=selectFolderPath3)
                image_button3.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame4 = tk.Frame(windowImg)
                image_frame4.grid(row=3, column=0, pady=2)

                image_label4 = tk.Label(image_frame4, text="Test Data path 4")
                image_label4.grid(row=0, column=0, ipadx=10)
                image_entry4 = tk.Entry(image_frame4, textvariable=folderPath4)
                image_entry4.grid(row=0, column=1, ipadx=124)
                image_button4 = tk.Button(image_frame4, text="Select", width=6, command=selectFolderPath4)
                image_button4.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame5 = tk.Frame(windowImg)
                image_frame5.grid(row=4, column=0, pady=2)

                image_label5 = tk.Label(image_frame5, text="Test Data path 5")
                image_label5.grid(row=0, column=0, ipadx=10)
                image_entry5 = tk.Entry(image_frame5, textvariable=folderPath5)
                image_entry5.grid(row=0, column=1, ipadx=124)
                image_button5 = tk.Button(image_frame5, text="Select", width=6, command=selectFolderPath5)
                image_button5.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame6 = tk.Frame(windowImg)
                image_frame6.grid(row=5, column=0, pady=2)

                image_label6 = tk.Label(image_frame6, text="Test Data path 6")
                image_label6.grid(row=0, column=0, ipadx=10)
                image_entry6 = tk.Entry(image_frame6, textvariable=folderPath6)
                image_entry6.grid(row=0, column=1, ipadx=124)
                image_button6 = tk.Button(image_frame6, text="Select", width=6, command=selectFolderPath6)
                image_button6.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame7 = tk.Frame(windowImg)
                image_frame7.grid(row=6, column=0, pady=2)

                image_label7 = tk.Label(image_frame7, text="Test Data path 7")
                image_label7.grid(row=0, column=0, ipadx=10)
                image_entry7 = tk.Entry(image_frame7, textvariable=folderPath7)
                image_entry7.grid(row=0, column=1, ipadx=124)
                image_button7 = tk.Button(image_frame7, text="Select", width=6, command=selectFolderPath7)
                image_button7.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame8 = tk.Frame(windowImg)
                image_frame8.grid(row=7, column=0, pady=2)

                image_label8 = tk.Label(image_frame8, text="Test Data path 8")
                image_label8.grid(row=0, column=0, ipadx=10)
                image_entry8 = tk.Entry(image_frame8, textvariable=folderPath8)
                image_entry8.grid(row=0, column=1, ipadx=124)
                image_button8 = tk.Button(image_frame8, text="Select", width=6, command=selectFolderPath8)
                image_button8.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                image_frame9 = tk.Frame(windowImg)
                image_frame9.grid(row=8, column=0, pady=2)

                image_label9 = tk.Label(image_frame9, text="Test Data path 9")
                image_label9.grid(row=0, column=0, ipadx=10)
                image_entry9 = tk.Entry(image_frame9, textvariable=folderPath9)
                image_entry9.grid(row=0, column=1, ipadx=124)
                image_button9 = tk.Button(image_frame9, text="Select", width=6, command=selectFolderPath9)
                image_button9.grid(row=0, column=2, padx=7, pady=4)

                # ==================== #

                button_frameTD = tk.Frame(windowImg)
                button_frameTD.grid(row=9, column=0)

                start_buttonTD = tk.Button(button_frameTD, text="Confirm", width=9, command=windowImg.destroy)
                start_buttonTD.grid(row=0, column=0, padx=10, pady=10)
                exit_buttonTD = tk.Button(button_frameTD, text="Close", width=9, command=exit_)
                exit_buttonTD.grid(row=0, column=1, padx=10, pady=10)

                windowWB.mainloop()
                folderPath = []

                print("Test Data path is : \n")
                if folderPath1.get() != "":
                    print(folderPath1.get(), "\n")
                    folderPath.append(folderPath1.get())
                if folderPath2.get() != "":
                    print(folderPath2.get(), "\n")
                    folderPath.append(folderPath2.get())
                if folderPath3.get() != "":
                    print(folderPath3.get(), "\n")
                    folderPath.append(folderPath3.get())
                if folderPath4.get() != "":
                    print(folderPath4.get(), "\n")
                    folderPath.append(folderPath4.get())
                if folderPath5.get() != "":
                    print(folderPath5.get(), "\n")
                    folderPath.append(folderPath5.get())
                if folderPath6.get() != "":
                    print(folderPath6.get(), "\n")
                    folderPath.append(folderPath6.get())
                if folderPath7.get() != "":
                    print(folderPath7.get(), "\n")
                    folderPath.append(folderPath7.get())
                if folderPath8.get() != "":
                    print(folderPath8.get(), "\n")
                    folderPath.append(folderPath8.get())
                if folderPath9.get() != "":
                    print(folderPath9.get(), "\n")
                    folderPath.append(folderPath9.get())
                print("========================================\n")

    for executeCount in range(len(limitFilePath)):

        # ================================================================================ #
        #                                                                                  #
        #                        按照 Limit File 檢查或編列測項 Excel                        #
        #                                                                                  #
        # ================================================================================ #

        if True:

            if excelPath[-5:] != ".xlsx":
                newExeclName = "DVT test report_" + datetime.now().strftime('%Y%m%d') + ".xlsx"
                shutil.copyfile(os.path.join((pathlib.Path().absolute()), "DVT test report.xlsx"), os.path.join(excelPath, newExeclName))
                excelPath = excelPath + "/" + newExeclName

            # 打開 Config
            configPathTxt = configPath + " Config.txt"
            txt = open(configPathTxt, "r")
            configArray = (txt.read()).split("\n")
            virtual = configArray.pop(0)
            for configList in range(len(configArray)):
                if "" in configArray:
                    configArray.remove("")

            # 存取 Excel 內所有工作表的名稱
            ExcelWB = openpyxl.load_workbook(excelPath)
            ExcelWS = ExcelWB.sheetnames

            # 存取 Limit File 內所有工作表的名稱
            LimitFileWB = openpyxl.load_workbook(limitFilePath[executeCount])
            LimitFileWS = LimitFileWB.sheetnames

            # 選定要編輯的工作表
            MappingWS = LimitFileWB["SheetNameMapping"]

            normal = ""
            low = ""
            high = ""

            # 定義路徑欄
            for row in MappingWS['A1':'A10']:
                for cell in row:
                    if cell.value == None:
                        break
                    if cell.value[-4:] == "+25C":
                        normal = MappingWS.cell(row=cell.row, column=2).value
                    if cell.value[-4:] == "-40C":
                        low = MappingWS.cell(row=cell.row, column=2).value
                    if cell.value[-4:] == "+55C":
                        high = MappingWS.cell(row=cell.row, column=2).value

            name = (limitFilePath[executeCount]).split("/")[-1][10:-5]

            if name in ExcelWS:
                editWS = ExcelWB[name]
            else:
                print("Building test item by the limit file\n\n========================================\n")
                Example = ExcelWB["Example"]
                editWS = ExcelWB.copy_worksheet(Example)
                editWS.title = name

                if normal != "":
                    NormalWS = LimitFileWB[normal]
                    for NormalWsCount in range(1, NormalWS.max_row+1):
                        TS_No = ("TS_#%.3d" % NormalWsCount)
                        Temp = "+25C"
                        Test_Name = NormalWS.cell(row=NormalWsCount, column=7).value
                        Unit = NormalWS.cell(row=NormalWsCount, column=6).value
                        Compare = NormalWS.cell(row=NormalWsCount, column=4).value
                        L_LMT = NormalWS.cell(row=NormalWsCount, column=3).value
                        H_LMT = NormalWS.cell(row=NormalWsCount, column=2).value

                        editWS.cell(row=NormalWsCount+9, column=1).value = TS_No
                        editWS.cell(row=NormalWsCount+9, column=3).value = Temp
                        editWS.cell(row=NormalWsCount+9, column=4).value = Test_Name
                        editWS.cell(row=NormalWsCount+9, column=7).value = Unit
                        editWS.cell(row=NormalWsCount+9, column=8).value = Compare
                        editWS.cell(row=NormalWsCount+9, column=9).value = L_LMT
                        editWS.cell(row=NormalWsCount+9, column=10).value = H_LMT

                        Counter = NormalWsCount

                if low != "":
                    LowWS = LimitFileWB[low]
                    for LowWsCount in range(1, LowWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+LowWsCount))
                        Temp = "-40C"
                        Test_Name = LowWS.cell(row=LowWsCount, column=7).value
                        Unit = LowWS.cell(row=LowWsCount, column=6).value
                        Compare = LowWS.cell(row=LowWsCount, column=4).value
                        L_LMT = LowWS.cell(row=LowWsCount, column=3).value
                        H_LMT = LowWS.cell(row=LowWsCount, column=2).value

                        editWS.cell(row=Counter+LowWsCount+9, column=1).value = TS_No
                        editWS.cell(row=Counter+LowWsCount+9, column=3).value = Temp
                        editWS.cell(row=Counter+LowWsCount+9, column=4).value = Test_Name
                        editWS.cell(row=Counter+LowWsCount+9, column=7).value = Unit
                        editWS.cell(row=Counter+LowWsCount+9, column=8).value = Compare
                        editWS.cell(row=Counter+LowWsCount+9, column=9).value = L_LMT
                        editWS.cell(row=Counter+LowWsCount+9, column=10).value = H_LMT

                        LowCount = LowWsCount

                    Counter = Counter + LowCount

                if high != "":
                    HighWS = LimitFileWB[high]
                    for HighWsCount in range(1, HighWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+HighWsCount))
                        Temp = "+55C"
                        Test_Name = HighWS.cell(row=HighWsCount, column=7).value
                        Unit = HighWS.cell(row=HighWsCount, column=6).value
                        Compare = HighWS.cell(row=HighWsCount, column=4).value
                        L_LMT = HighWS.cell(row=HighWsCount, column=3).value
                        H_LMT = HighWS.cell(row=HighWsCount, column=2).value

                        editWS.cell(row=Counter+HighWsCount+9, column=1).value = TS_No
                        editWS.cell(row=Counter+HighWsCount+9, column=3).value = Temp
                        editWS.cell(row=Counter+HighWsCount+9, column=4).value = Test_Name
                        editWS.cell(row=Counter+HighWsCount+9, column=7).value = Unit
                        editWS.cell(row=Counter+HighWsCount+9, column=8).value = Compare
                        editWS.cell(row=Counter+HighWsCount+9, column=9).value = L_LMT
                        editWS.cell(row=Counter+HighWsCount+9, column=10).value = H_LMT

                        HighCount = HighWsCount

                    Counter = Counter + HighCount

        # ================================================================================ #
        #                                                                                  #
        #                                     前置作業                                      #
        #                                                                                  #
        # ================================================================================ #

        if True:

            # 要編輯的工作表
            # editWS

            unitCoordinate = 2
            validCoordinate = "novalid"
            aCoordinate     = "noa"
            bCoordinate     = "nob"
            cCoordinate     = "noc"
            dCoordinate     = "nod"
            eCoordinate     = "noe"
            fCoordinate     = "nof"
            gCoordinate     = "nog"
            hCoordinate     = "noh"
            iCoordinate     = "noi"
            jCoordinate     = "noj"
            kCoordinate     = "nok"
            lCoordinate     = "nol"
            mCoordinate     = "nom"
            nCoordinate     = "non"
            # 定義路徑欄
            for row in editWS['A9':'AZ9']:
                for cell in row:
                    if cell.value == "Temp.":
                        tempCoordinate = cell.column
                    if cell.value == "Test_Name":
                        nameCoordinate = cell.column
                    if cell.value == "Result":
                        resultCoordinate = cell.column
                    if cell.value == "TestVal":
                        valueCoordinate = cell.column
                    if cell.value == "Image_Local":
                        imageLocalCoordinate = cell.column
                    if cell.value == "Image_Server":
                        imageServerCoordinate = cell.column
                    if cell.value == "Session":
                        sessionCoordinate = cell.column
                    if cell.value == "Start_Time":
                        timeCoordinate = cell.column
                    if cell.value == "Exec_Time":
                        execCoordinate = cell.column
                    if cell.value == "A_TestParam":
                        aCoordinate = cell.column
                    if cell.value == "B_TestParam":
                        bCoordinate = cell.column
                    if cell.value == "C_TestParam":
                        cCoordinate = cell.column
                    if cell.value == "D_TestParam":
                        dCoordinate = cell.column
                    if cell.value == "E_TestParam":
                        eCoordinate = cell.column
                    if cell.value == "F_TestParam":
                        fCoordinate = cell.column
                    if cell.value == "G_TestParam":
                        gCoordinate = cell.column
                    if cell.value == "H_TestParam":
                        hCoordinate = cell.column
                    if cell.value == "I_TestParam":
                        iCoordinate = cell.column
                    if cell.value == "J_TestParam":
                        jCoordinate = cell.column
                    if cell.value == "K_TestParam":
                        kCoordinate = cell.column
                    if cell.value == "L_TestParam":
                        lCoordinate = cell.column
                    if cell.value == "M_TestParam":
                        mCoordinate = cell.column
                    if cell.value == "N_TestParam":
                        nCoordinate = cell.column
                    if cell.value == "Validation":
                        validCoordinate = cell.column
                    if cell.value == "Test Plot":
                        picCoordinate = cell.column

            # 將所有符合工作表名稱的資料夾放入 folder[]
            folder = []

            for initialFolderPathCount in range(len(folderPath)):
                dir = ([ name for name in os.listdir(folderPath[initialFolderPathCount]) if os.path.isdir(os.path.join(folderPath[initialFolderPathCount], name)) ])
                for initialFolderDir in range(len(dir)):
                    if name == dir[initialFolderDir]:  folder.append(folderPath[initialFolderPathCount] + "/" + dir[initialFolderDir])

            # 列出所有不同時間、不同溫度之 Test Log
            testLogFolder = []

            for testLogFolderCount in range(len(folder)):
                dir = ([ name for name in os.listdir(folder[testLogFolderCount]) if os.path.isdir(os.path.join(folder[testLogFolderCount], name)) ])
                for testLogFolderDir in range(len(dir)):
                    if len(dir[testLogFolderDir].split("][")) != 5:
                        continue
                    testLogFolder.append(folder[testLogFolderCount] + "/" + dir[testLogFolderDir])

            testLogFolder.reverse()
            # 將 Test Log 依時間排序
            folderTimeStr = [""] * len(testLogFolder)
            folderTimeSec = [""] * len(testLogFolder)
            for testLogFolderCount in range(len(testLogFolder)):
                folderTimeStr[testLogFolderCount] = (testLogFolder[testLogFolderCount].split("][")[2]) + "_" + (testLogFolder[testLogFolderCount].split("][")[3])
                folderTimeSec[testLogFolderCount] = time.mktime((datetime.strptime(folderTimeStr[testLogFolderCount], "%Y-%m-%d_%H_%M_%S")).timetuple())

            testLogFolderTime = list(zip(testLogFolder, folderTimeSec))
            sortTestLogFolder = sorted(testLogFolderTime, key= lambda testLogFolderTime : testLogFolderTime[1])
            sortTestLogFolder = np.array(sortTestLogFolder)[:, 0]

        # ================================================================================ #
        #                                                                                  #
        #                                     掃描測項                                      #
        #                                                                                  #
        # ================================================================================ #

        print("Scan file, fill in Test Data and copy image . . . \n")
        print("Running : %s\n" % name)
        startTime = time.time()
        print(time.strftime("Start time : %Y/%m/%d %a %H:%M:%S\n", time.localtime()))

        # 宣告最後用來放入 Pic 超連結路徑的 Array
        picPath = ["Path"] * (editWS.max_row + 1 - 10)
        picName = ["Name"] * (editWS.max_row + 1 - 10)
        testPlot = ["Plot"] * (editWS.max_row + 1 - 10)

        temp0 = 0
        temp1 = 0
        total0 = editWS.max_row-9
        picDir1 = []
        picDir2 = []
        execTime = time.time()
        ftpFolderLink = "\\\\twhsi-pub10.hq.mti.inc\\Publish_Data\\Eng_" + virtual
        newFTPFolderTestItemUserPath = ""

        @retry_on_communication_error
        def TestDataBuilding(rowCount, count):

            global temp0, picDir1, picDir2, execTime, ftpFolderLink, newFTPFolderTestItemUserPath, picPath, picName, testPlot
            temp0 = rowCount-9
            picDir1 = []
            picDir2 = []

            result = []
            valid  = []

            for resultCount in range(count):
                # 儲存結果欄位
                result.append(editWS.cell(row=rowCount+resultCount, column=resultCoordinate).value)
            if validCoordinate != "novalid":
                for validCount in range(count):
                    # 儲存結果欄位
                    valid.append(editWS.cell(row=rowCount+validCount, column=validCoordinate).value)

            if (None in result) or ("Failed" in result) or ("Skipped" in result) or ("Invalid" in valid):

                testName = []
                newTestName = [""] * count
                compare = 0.5

                for groupingCount in range(count):

                    # 儲存名稱欄位
                    testName.append(editWS.cell(row=rowCount+groupingCount, column=nameCoordinate).value)

                    # 去掉 Test Name 前面的項目名，只留下後段測試名
                    nameSplit = testName[groupingCount].split("][")
                    newTestName[groupingCount] = "[" + (nameSplit[0].split("["))[1]
                    for nameSplitCount in range(1, len(nameSplit)):
                        newTestName[groupingCount] = newTestName[groupingCount] + "][" + nameSplit[nameSplitCount]

                # 儲存溫度欄位
                temp = (editWS.cell(row=rowCount, column=tempCoordinate).value)

                for sortTestLogFolderCount in range(len(sortTestLogFolder)):
                    # 篩選掉溫度不符合的
                    if temp == (sortTestLogFolder[sortTestLogFolderCount].split("/")[-1]).split("][")[1]:
                        testDataDir = os.listdir(sortTestLogFolder[sortTestLogFolderCount])
                        testDataDir.sort()
                        for testPicDirCount in range(len(testDataDir)):
                            if name == testDataDir[testPicDirCount]:
                                picFolder1 = (sortTestLogFolder[sortTestLogFolderCount] + "/" + testDataDir[testPicDirCount])
                                picDir1 = os.listdir(picFolder1)
                            if testDataDir[testPicDirCount][-5:] == "_PASS" or testDataDir[testPicDirCount][-5:] == "_FAIL":
                                if name == testDataDir[testPicDirCount][:-5]:
                                    picFolder2 = (sortTestLogFolder[sortTestLogFolderCount] + "/" + testDataDir[testPicDirCount])
                                    picDir2 = os.listdir(picFolder2)
                        for testDataDirCount in range(len(testDataDir)):
                            if testDataDir[testDataDirCount][-16:] == "_TestReport.xlsx" and testDataDir[testDataDirCount][:1] == "(" and testDataDir[testDataDirCount][:2] != "~$":
                                finalFolder = (sortTestLogFolder[sortTestLogFolderCount] + "/" + testDataDir[testDataDirCount])
                                TestDataWB = openpyxl.load_workbook(finalFolder)
                                if name not in TestDataWB.sheetnames:
                                    continue
                                TestDataWS = TestDataWB[name]

                                for TestDataWsCount in range(12, TestDataWS.max_row+1):
                                    # Grouping
                                    trueCounter = 0
                                    for groupingCount in range(count):
                                        if editWS.cell(row=rowCount+groupingCount, column=8).value == TestDataWS.cell(row=TestDataWsCount+groupingCount, column=7).value:
                                            if testName[groupingCount] == TestDataWS.cell(row=TestDataWsCount+groupingCount, column=3).value and TestDataWS.cell(row=TestDataWsCount+groupingCount, column=4).value == "Passed":
                                                trueCounter += 4
                                            if testName[groupingCount] == TestDataWS.cell(row=TestDataWsCount+groupingCount, column=3).value and TestDataWS.cell(row=TestDataWsCount+groupingCount, column=4).value == "Failed":
                                                trueCounter += 3
                                            if testName[groupingCount] == TestDataWS.cell(row=TestDataWsCount+groupingCount, column=3).value and TestDataWS.cell(row=TestDataWsCount+groupingCount, column=4).value == "Skipped":
                                                trueCounter += 2
                                            if testName[groupingCount] == TestDataWS.cell(row=TestDataWsCount+groupingCount, column=3).value and TestDataWS.cell(row=TestDataWsCount+groupingCount, column=4).value == None:
                                                trueCounter += 1

                                    if trueCounter <= count*4 and trueCounter >= compare:
                                        compare = trueCounter

                                        for groupingCount in range(count):
                                            editWS.cell(row=rowCount+groupingCount, column=unitCoordinate).value = (testDataDir[testDataDirCount][1:-1]).split(")(")[0]                 #dataUnit
                                            editWS.cell(row=rowCount+groupingCount, column=resultCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=4).value #dataResult
                                            editWS.cell(row=rowCount+groupingCount, column=valueCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=5).value  #dataTestVal
                                            editWS.cell(row=rowCount+groupingCount, column=sessionCoordinate).value = finalFolder.split("/")[-2]                                        #dataSession
                                            editWS.cell(row=rowCount+groupingCount, column=timeCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=11).value  #dataStart
                                            editWS.cell(row=rowCount+groupingCount, column=execCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=12).value  #dataExec
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=13).value != "" and aCoordinate != "noa":
                                                editWS.cell(row=rowCount+groupingCount, column=aCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=13).value #dataA
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=14).value != "" and bCoordinate != "nob":
                                                editWS.cell(row=rowCount+groupingCount, column=bCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=14).value #dataB
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=15).value != "" and cCoordinate != "noc":
                                                editWS.cell(row=rowCount+groupingCount, column=cCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=15).value #dataC
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=16).value != "" and dCoordinate != "nod":
                                                editWS.cell(row=rowCount+groupingCount, column=dCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=16).value #dataD
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=17).value != "" and eCoordinate != "noe":
                                                editWS.cell(row=rowCount+groupingCount, column=eCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=17).value #dataE
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=18).value != "" and fCoordinate != "nof":
                                                editWS.cell(row=rowCount+groupingCount, column=fCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=18).value #dataF
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=19).value != "" and gCoordinate != "nog":
                                                editWS.cell(row=rowCount+groupingCount, column=gCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=19).value #dataG
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=20).value != "" and hCoordinate != "noh":
                                                editWS.cell(row=rowCount+groupingCount, column=hCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=20).value #dataH
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=21).value != "" and iCoordinate != "noi":
                                                editWS.cell(row=rowCount+groupingCount, column=iCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=21).value #dataI
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=22).value != "" and jCoordinate != "noj":
                                                editWS.cell(row=rowCount+groupingCount, column=jCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=22).value #dataJ
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=23).value != "" and kCoordinate != "nok":
                                                editWS.cell(row=rowCount+groupingCount, column=kCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=23).value #dataK
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=24).value != "" and lCoordinate != "nol":
                                                editWS.cell(row=rowCount+groupingCount, column=lCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=24).value #dataL
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=25).value != "" and mCoordinate != "nom":
                                                editWS.cell(row=rowCount+groupingCount, column=mCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=25).value #dataM
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=26).value != "" and nCoordinate != "non":
                                                editWS.cell(row=rowCount+groupingCount, column=nCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=26).value #dataN
                                            if validCoordinate != "novalid":
                                                editWS.cell(row=rowCount+groupingCount, column=validCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=27).value #dataValidation
                                            if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value != None:
                                                if TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value[-4:] == ".png":
                                                    editWS.cell(row=rowCount+groupingCount, column=picCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value[:-4] #dataTestPlot
                                                    if (("NEC" in configPathText.get()) or ("TB" in configPathText.get()) or ("Quadriga" in configPathText.get())) and "RX_NF" in name:
                                                        testPlot[(rowCount+groupingCount)-10] = (TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value).split(",")[1]
                                                    else:
                                                        testPlot[(rowCount+groupingCount)-10] = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value
                                                else:
                                                    editWS.cell(row=rowCount+groupingCount, column=picCoordinate).value = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value #dataTestPlot
                                                    if (("NEC" in configPathText.get()) or ("TB" in configPathText.get()) or ("Quadriga" in configPathText.get())) and "RX_NF" in name:
                                                        testPlot[(rowCount+groupingCount)-10] = (TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value).split(",")[1] + ".png"
                                                    else:
                                                        testPlot[(rowCount+groupingCount)-10] = TestDataWS.cell(row=TestDataWsCount+groupingCount, column=29).value + ".png"

                                                # 搜尋符合的照片，記錄其所在資料夾以及名稱
                                                for picDir1Count in range(len(picDir1)):
                                                    if picDir1[picDir1Count] == testPlot[(rowCount+groupingCount)-10]:
                                                        picPath[(rowCount+groupingCount)-10] = picFolder1
                                                        picName[(rowCount+groupingCount)-10] = newTestName[groupingCount] + "_" + (testDataDir[testDataDirCount][1:-1]).split(")(")[0][-4:] + ".png"

                                                for picDir2Count in range(len(picDir2)):
                                                    if picDir2[picDir2Count] == testPlot[(rowCount+groupingCount)-10]:
                                                        picPath[(rowCount+groupingCount)-10] = picFolder2
                                                        picName[(rowCount+groupingCount)-10] = newTestName[groupingCount] + "_" + (testDataDir[testDataDirCount][1:-1]).split(")(")[0][-4:] + ".png"

                for groupingCount in range(count):
                    if picPath[rowCount+groupingCount-10] != "Path" and picName[rowCount+groupingCount-10] != "Name":
                        # 避免相同的 Test Plot 因 Test Name 不同，而重複複製成兩張不同的照片
                        if testPlot[rowCount+groupingCount-10] == testPlot[rowCount+groupingCount-11] and (rowCount+groupingCount) > 10:
                            picName[rowCount+groupingCount-10] = picName[rowCount+groupingCount-11]

                        localNewPath = "Image/" + newImageFolder.split("/")[-1] + "/" + picName[rowCount+groupingCount-10]
                        serverNewPath = newFTPFolderTestItemUserPath + "/" + picName[rowCount+groupingCount-10]
                        editWS.cell(row=rowCount+groupingCount, column=imageLocalCoordinate).value = ('=HYPERLINK("%s", "Link")' % localNewPath)
                        editWS.cell(row=rowCount+groupingCount, column=imageLocalCoordinate).font = Font(underline="single", color='00B050')
                        editWS.cell(row=rowCount+groupingCount, column=imageServerCoordinate).value = ('=HYPERLINK("%s", "Remote_Link")' % serverNewPath)
                        editWS.cell(row=rowCount+groupingCount, column=imageServerCoordinate).font = Font(underline="single", color='00B050')
                        try:
                            shutil.copyfile(os.path.join(picPath[rowCount+groupingCount-10], testPlot[rowCount+groupingCount-10]),os.path.join(newImageFolder, picName[rowCount+groupingCount-10]))
                            shutil.copyfile(os.path.join(picPath[rowCount+groupingCount-10], testPlot[rowCount+groupingCount-10]),os.path.join(newFTPFolderTestItemPath, picName[rowCount+groupingCount-10]))
                        except:
                            print("Image copy Error.\n")
                            continue

                    """
                    if (rowCount+groupingCount-10) % 2000 == 0 and i > 10:
                        print("\n\nSaving!", end="")
                        executeCount = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - execTime))
                        print("     Execute time: %d Hr %d Min %d Sec\n" % (int(executeCount.split("/")[3])-8, int(executeCount.split("/")[4]), int(executeCount.split("/")[5])))
                        execTime = time.time()
                        ExcelWB.save(excelPath.get())
                    """

                    # Progress Bar ( 掃描檔案，填入 Test Data )
                    print('\r' + '%.2f%% |%s%s| %d/%d' % ((float((temp0+groupingCount)/total0*100)), '█' * int((temp0+groupingCount)*50/total0), ' ' * (50-int((temp0+groupingCount)*50/total0)), (rowCount+groupingCount-9), (editWS.max_row-9)), end='')

        #@retry_on_communication_error
        """
        def TestImageUpdate(rowCount, count):
            global temp1, picPath, picName, testPlot
            temp1 = rowCount-9
            for groupingCount in range(count):
                try:
                    shutil.copyfile(os.path.join(picPath[rowCount+groupingCount-10], testPlot[rowCount+groupingCount-10]),os.path.join(newImageFolder, picName[rowCount+groupingCount-10]))
                    shutil.copyfile(os.path.join(picPath[rowCount+groupingCount-10], testPlot[rowCount+groupingCount-10]),os.path.join(newFTPFolderTestItemPath, picName[rowCount+groupingCount-10]))
                except:
                    print("Image copy Error.\n")
                    continue

                # Progress Bar ( 掃描檔案，填入 Test Data )
                print('\r' + '%.2f%% |%s%s| %d/%d' % ((float((temp1+groupingCount)/total0*100)), '█' * int((temp1+groupingCount)*50/total0), ' ' * (50-int((temp1+groupingCount)*50/total0)), (rowCount+groupingCount-9), (editWS.max_row-9)), end='')
        """

        # Grouping Count
        if True:
            for group in range(len(configArray)):
                if (configArray[group].split(","))[0] in name:
                    groupCount = int((configArray[group].split(","))[1])

        # Create Image Folder
        if True:
            # 查看 Image 資料夾是否存在，若無，建立 Image 資料夾
            newImagePath = os.path.dirname(excelPath)
            if not(os.path.exists(newImagePath + "/Image")):
                os.makedirs(newImagePath + "/Image")

            # 查看測項子 Image 資料夾是否存在，若無，建立測項子 Image 資料夾
            newImageFolder = newImagePath + "/Image/" + name
            if not(os.path.exists(newImageFolder)):
                os.makedirs(newImageFolder)

            # 查看 Image 資料夾是否存在於 FTP Folder，若無，建立 Image 資料夾
            newFTPFolderImagePath = ([ name for name in os.listdir(ftpFolderLink) if os.path.isdir(os.path.join(ftpFolderLink, name)) ])

            # 查看測試 Model 資料夾是否存在於 FTP Folder，若無，建立測試 Model 資料夾
            newFTPFolderModelPath = ftpFolderLink + "/Image/" + configPath
            if not(os.path.exists(newFTPFolderModelPath)):
                os.makedirs(newFTPFolderModelPath)

            # 查看測試 Phase 資料夾是否存在於 Model，若無，建立測試 Phase 資料夾
            newFTPFolderPhasePath = newFTPFolderModelPath + "/" + phaseSelect
            if not(os.path.exists(newFTPFolderPhasePath)):
                os.makedirs(newFTPFolderPhasePath)

            # 查看測試 Test Item 資料夾是否存在於 Phase，若無，建立測試 Test Item 資料夾
            newFTPFolderTestItemPath = newFTPFolderPhasePath + "/" + name
            if not(os.path.exists(newFTPFolderTestItemPath)):
                os.makedirs(newFTPFolderTestItemPath)

            ftpFolderUserLink = "https://twpub10.mtigroup.com/" + virtual + "/"
            newFTPFolderTestItemUserPath = ftpFolderUserLink + "Image/" + configPath + "/" + phaseSelect + "/" + name

        for rowCounter in range(10, editWS.max_row+1, groupCount):
            scanErrorCount = 0
            while True:
                try:
                    TestDataBuilding(rowCounter, groupCount)
                    break
                except:
                    print("\n\nData searching Error.  Retrying.\n")
                    scanErrorCount += 1
                    if scanErrorCount == 5:
                        print("Timeout, program ends\n")
                        print("Please check Limit File, Test Data, network connection, whether any item is wrong\n")
                        time.sleep(5)
                        pause()
                    time.sleep(2)
                    #continue

        # ================================================================================ #
        #                                                                                  #
        #                                檔案儲存 ( Excel )                                 #
        #                                                                                  #
        # ================================================================================ #

        if True:

            # Test Case Title
            editWS.cell(row=3, column=2).value = name
            editWS.cell(row=3, column=2).font = Font(name='Calibri', size=9)

            # Sheet Result Summary
            resultRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",E3)))']
            editWS.conditional_formatting.add('E3', resultRulePassed)
            resultRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",E3)))']
            editWS.conditional_formatting.add('E3', resultRuleFailed)

            editWS.cell(row=4, column=5).value = ('=SUBTOTAL(103,$E$10:$E$%d)' % editWS.max_row) # E4
            editWS.cell(row=6, column=5).value = ('=SUMPRODUCT(SUBTOTAL(3,OFFSET($E$10:$E$%d,ROW($E$10:$E$%d)-MIN(ROW($E$10:$E$%d)),,1))*($E$10:$E$%d="Passed"))' % (editWS.max_row, editWS.max_row, editWS.max_row, editWS.max_row)) # E6

            # Sheet Test Item Result
            dataCell = 'E10:E%d' % (editWS.max_row)
            dataRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            dataRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",E10)))']
            editWS.conditional_formatting.add(dataCell, dataRulePassed)
            dataRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            dataRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",E10)))']
            editWS.conditional_formatting.add(dataCell, dataRuleFailed)
            dataRuleSkipped = Rule(type="containsText", text="Skipped", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FFFF00'), alignment=Alignment(horizontal='center', vertical='bottom')))
            dataRuleSkipped.formula = ['NOT(ISERROR(SEARCH("Skipped",E10)))']
            editWS.conditional_formatting.add(dataCell, dataRuleSkipped)

            # Results Overview
            resultWS = ExcelWB["Results_Overview"]
            resultWS.cell(row=resultWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            resultWS.cell(row=resultWS.max_row, column=1).font = Font(underline="single", color='0563C1')
            resultWS.cell(row=resultWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            resultOverviewRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultOverviewRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",B12)))']
            resultWS.conditional_formatting.add(resultWS.cell(row=resultWS.max_row, column=2).coordinate, resultOverviewRulePassed)
            resultOverviewRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultOverviewRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",B12)))']
            resultWS.conditional_formatting.add(resultWS.cell(row=resultWS.max_row, column=2).coordinate, resultOverviewRuleFailed)
            resultOverviewRuleIncomplete = Rule(type="containsText", text="Incomplete", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FFFF00'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultOverviewRuleIncomplete.formula = ['NOT(ISERROR(SEARCH("Incomplete",B12)))']
            resultWS.conditional_formatting.add(resultWS.cell(row=resultWS.max_row, column=2).coordinate, resultOverviewRuleIncomplete)
            resultOverviewRuleNoData = Rule(type="containsText", text="No Data", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FFFF00'), alignment=Alignment(horizontal='center', vertical='bottom')))
            resultOverviewRuleNoData.formula = ['NOT(ISERROR(SEARCH("No Data",B12)))']
            resultWS.conditional_formatting.add(resultWS.cell(row=resultWS.max_row, column=2).coordinate, resultOverviewRuleNoData)
            resultWS.cell(row=resultWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"-40C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+25C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+55C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))

            # Coverage Overview
            coverageWS = ExcelWB["Coverage_Overview"]
            coverageWS.cell(row=coverageWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            coverageWS.cell(row=coverageWS.max_row, column=1).font = Font(underline="single", color='0563C1')
            coverageWS.cell(row=coverageWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            coverageRulePassed = Rule(type="containsText", text="Passed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='00B050'), alignment=Alignment(horizontal='center', vertical='bottom')))
            coverageRulePassed.formula = ['NOT(ISERROR(SEARCH("Passed",B12)))']
            coverageWS.conditional_formatting.add(coverageWS.cell(row=coverageWS.max_row, column=2).coordinate, coverageRulePassed)
            coverageRuleFailed = Rule(type="containsText", text="Failed", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FF0000'), alignment=Alignment(horizontal='center', vertical='bottom')))
            coverageRuleFailed.formula = ['NOT(ISERROR(SEARCH("Failed",B12)))']
            coverageWS.conditional_formatting.add(coverageWS.cell(row=coverageWS.max_row, column=2).coordinate, coverageRuleFailed)
            coverageRuleIncomplete = Rule(type="containsText", text="Incomplete", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FFFF00'), alignment=Alignment(horizontal='center', vertical='bottom')))
            coverageRuleIncomplete.formula = ['NOT(ISERROR(SEARCH("Incomplete",B12)))']
            coverageWS.conditional_formatting.add(coverageWS.cell(row=coverageWS.max_row, column=2).coordinate, coverageRuleIncomplete)
            coverageRuleNoData = Rule(type="containsText", text="No Data", dxf=DifferentialStyle(fill=PatternFill("solid", bgColor='FFFF00'), alignment=Alignment(horizontal='center', vertical='bottom')))
            coverageRuleNoData.formula = ['NOT(ISERROR(SEARCH("No Data",B12)))']
            coverageWS.conditional_formatting.add(coverageWS.cell(row=coverageWS.max_row, column=2).coordinate, coverageRuleNoData)
            coverageWS.cell(row=coverageWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))

            print("\n\nSaving...  Please Wait...\n")
            print("========================================\n")
            #wb.save(excelPath.get()[:-5] + "_copy.xlsx")
            saveErrorCount = 0
            while True:
                try:
                    ExcelWB.save(excelPath)
                    break
                except:
                    print("\n\nConnection Error. Retry Saving.\n")
                    saveErrorCount += 1
                    if saveErrorCount == 5:
                        print("Timeout, program ends, Saving is not completed")
                        time.sleep(5)
                        pause()
                    time.sleep(2)
            root = tk.Tk()
            root.title("")
            root.geometry("170x70")
            root.resizable(0,0)
            frame1 = tk.Frame(root)
            frame1.grid(row=0, column=0, pady=2)

            label1 = tk.Label(frame1, text="Report Generate success")
            label1.grid(row=0, column=0, ipadx=10)
            button1 = tk.Button(frame1, text="Continue", width=6, command=root.destroy)
            button1.grid(row=1, column=0, padx=20)
            root.after(20000, root.destroy)
            root.mainloop()
            print("Success!\n")
            blockTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - startTime))
            totalTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(time.time() - totalStartTime))
            print("Block execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(blockTime.split("/")[2])-1, int(blockTime.split("/")[3])-8, int(blockTime.split("/")[4]), int(blockTime.split("/")[5])))
            print("Total execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(totalTime.split("/")[2])-1, int(totalTime.split("/")[3])-8, int(totalTime.split("/")[4]), int(totalTime.split("/")[5])))
            print("========================================\n")
            time.sleep(2)

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    if True:

        print("Re-run Y/N ?")
        executeYN = input()
        print()
        print("========================================\n")

        if executeYN == "Y" or executeYN == "y":
            execute = True
        elif executeYN == "N" or executeYN == "n":
            execute = False
        if execute == False:
            break
