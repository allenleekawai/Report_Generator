# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import cv2 as cv

import openpyxl

import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename

# 宣告 "關閉" 按鈕
def exit_():
    sys.exit()

def selectExcelPath():
    excelPath_ = askopenfilename()
    excelPath.set(excelPath_)

if True: # 建立選擇 Excel GUI

    print("請選擇 Excel 路徑\n")

    # 建立使用者 GUI，使其可選擇 Excel 路徑
    windowWB = tk.Tk()
    windowWB.title("Excel 路徑選擇")
    windowWB.geometry('560x120')
    windowWB.resizable(0,0)

    excelPath = StringVar()

    excel_frame2 = tk.Frame(windowWB)
    excel_frame2.grid(row=0, column=0, pady=2)

    excel_label2 = tk.Label(excel_frame2, text="Excel 路徑")
    excel_label2.grid(row=0, column=0, ipadx=10)
    excel_entry2 = tk.Entry(excel_frame2, textvariable=excelPath)
    excel_entry2.grid(row=0, column=1, ipadx=124)
    excel_button2 = tk.Button(excel_frame2, text="選擇", command=selectExcelPath)
    excel_button2.grid(row=0, column=2, padx=7, pady=4)

    # ==================== #

    button_frameEX = tk.Frame(windowWB)
    button_frameEX.grid(row=1, column=0)

    start_buttonEX = tk.Button(button_frameEX, text="確認", width=6, command=windowWB.destroy)
    start_buttonEX.grid(row=0, column=0, padx=8)
    exit_buttonEX = tk.Button(button_frameEX, text="關閉", width=6, command=exit_)
    exit_buttonEX.grid(row=0, column=1, padx=8)

    windowWB.mainloop()

    print("Excel 路徑為：\n")
    if excelPath.get() != "":
        print(excelPath.get(), "\n")
    print("========================================\n")

#rightNowPath = os.path.join(pathlib.Path().absolute())

folder = os.path.abspath(os.path.join(excelPath.get(), os.path.pardir))
print(folder, end="\n\n")

ExcelWB = openpyxl.load_workbook(excelPath.get())
ExcelWS = ExcelWB.sheetnames

ExcelWS.remove("Overview")

print(ExcelWS, end="\n\n")

for name in ExcelWS:
    editWS = ExcelWB[name]
    for j in range(2, editWS.max_row+1):
        if (editWS.cell(row=j, column=6).value != None) and (editWS.cell(row=j, column=6).value != ""):
            pic = (editWS.cell(row=j, column=6).value)[12:-7]
            path = folder + "\\" + pic
            print(name, "     ", path)
            while True: 
                try:
                    img = cv.imread(path)
                    sh = img.shape
                    break
                except:
                    print("Error !!!!!!!!!!!!!!!")
                    exit_()
              