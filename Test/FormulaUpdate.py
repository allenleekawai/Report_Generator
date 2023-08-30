# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import pathlib

import openpyxl
from openpyxl.styles import Font

import time

import tkinter as tk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

execute = True

while True:

    # 宣告 "關閉" 按鈕
    def exit_():
        sys.exit()

    def selectExcelPath():
        excelPath_ = askopenfilename()
        excelPath.set(excelPath_)

    if True: # 建立選擇 Excel GUI

        print("請選擇 Excel 路徑\n")

        # 建立使用者 GUI，使其可選擇 Excel 路徑
        windowWB = tk.Tk()
        windowWB.title("Excel 路徑選擇")
        windowWB.geometry('560x120')
        windowWB.resizable(0,0)

        excelPath = StringVar()

        excel_frame2 = tk.Frame(windowWB)
        excel_frame2.grid(row=0, column=0, pady=2)

        excel_label2 = tk.Label(excel_frame2, text="Excel 路徑")
        excel_label2.grid(row=0, column=0, ipadx=10)
        excel_entry2 = tk.Entry(excel_frame2, textvariable=excelPath)
        excel_entry2.grid(row=0, column=1, ipadx=124)
        excel_button2 = tk.Button(excel_frame2, text="選擇", command=selectExcelPath)
        excel_button2.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        button_frameEX = tk.Frame(windowWB)
        button_frameEX.grid(row=1, column=0)

        start_buttonEX = tk.Button(button_frameEX, text="確認", width=6, command=windowWB.destroy)
        start_buttonEX.grid(row=0, column=0, padx=8)
        exit_buttonEX = tk.Button(button_frameEX, text="關閉", width=6, command=exit_)
        exit_buttonEX.grid(row=0, column=1, padx=8)

        windowWB.mainloop()

        print("Excel 路徑為：\n")
        if excelPath.get() != "":
            print(excelPath.get(), "\n")
        print("========================================\n")

    print("Updating formula...  Please Wait...\n")
    print("========================================\n")

    rightNowPath = os.path.join(pathlib.Path().absolute())

    ExcelWB = openpyxl.load_workbook(excelPath.get())

    resultWS = ExcelWB["Results_Overview"]
    for i in range(10, resultWS.max_row+1):
        resultWS.delete_rows(10)
    coverageWS = ExcelWB["Coverage_Overview"]
    for i in range(10, coverageWS.max_row+1):
        coverageWS.delete_rows(10)

    ExcelWS = ExcelWB.sheetnames
    CheckWS = ExcelWS

    CheckWS.remove("Results_Overview")
    CheckWS.remove("Coverage_Overview")
    CheckWS.remove("Example")

    for name in CheckWS:

        editWS = ExcelWB[name]

        # ================================================================================ #
        #                                                                                  #
        #                                檔案儲存 ( Excel )                                 #
        #                                                                                  #
        # ================================================================================ #

        if True:

            # Results Overview
            resultWS = ExcelWB["Results_Overview"]
            resultWS.cell(row=resultWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            resultWS.cell(row=resultWS.max_row, column=1).font = Font(underline="single", color='0563C1')
            resultWS.cell(row=resultWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            resultWS.cell(row=resultWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"-40C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+25C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            resultWS.cell(row=resultWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Failed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"",\'%s\'!C10:\'%s\'!C%d,"+55C")+COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Skipped",\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))

            # Coverage Overview
            coverageWS = ExcelWB["Coverage_Overview"]
            coverageWS.cell(row=coverageWS.max_row+1, column=1).value = ('=HYPERLINK("#%s!A1","%s")' % (name, name))
            coverageWS.cell(row=coverageWS.max_row, column=1).font = Font(underline="single", color='0563C1')
            coverageWS.cell(row=coverageWS.max_row, column=2).value = ('=\'%s\'!E3' % name)
            coverageWS.cell(row=coverageWS.max_row, column=4).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"-40C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"-40C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=5).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+25C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+25C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))
            coverageWS.cell(row=coverageWS.max_row, column=6).value = ('=COUNTIFS(\'%s\'!E10:\'%s\'!E%d,"Passed",\'%s\'!C10:\'%s\'!C%d,"+55C")&"/"&COUNTIF(\'%s\'!C10:\'%s\'!C%d,"+55C")' % (name, name, editWS.max_row, name, name, editWS.max_row, name, name, editWS.max_row))

    print("Saving...            Please Wait...\n")
    #wb.save(excelPath.get()[:-5] + "_copy.xlsx")
    saveErrorCount = 0
    while True:
        try:
            ExcelWB.save(excelPath.get())
            break
        except:
            print("Connection Error.    Retry Saving.\n")
            saveErrorCount += 1
            if saveErrorCount == 5:
                print("超過 Timeout 時間，程式結束，Saving 並未完成")
                time.sleep(5)
                exit_()
            time.sleep(2)
    print("Success!\n")
    print("========================================\n")
    time.sleep(2)

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    if True:

        print("Re-run Y/N ?")
        executeYN = input()
        print()
        print("========================================\n")

        if executeYN == "Y" or executeYN == "y":
            execute = True
        elif executeYN == "N" or executeYN == "n":
            execute = False
        if execute == False:
            break
