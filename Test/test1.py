
"""
import tkinter as tk
import tkinter.ttk as ttk
import sys

# 宣告 "關閉" 按鈕
def exit_():
    sys.exit()

windowWB = tk.Tk()
windowWB.title("功能選擇")
windowWB.geometry('180x140')
windowWB.resizable(0,0)

functionValue = tk.IntVar()

function_frame = tk.Frame(windowWB)
function_frame.grid(row=0, column=0, pady=2)

function_label = tk.Label(function_frame, text="功能選擇")
function_label.grid(row=0, column=0, ipadx=10, ipady=3, pady=5)
functionUpdate_check = ttk.Radiobutton(function_frame, text="Update Data", variable=functionValue, value=1)
functionUpdate_check.grid(row=1, column=0, ipadx=0, ipady=1, padx=40, sticky=tk.W)
functionNew_check = ttk.Radiobutton(function_frame, text="New Data", variable=functionValue, value=2)
functionNew_check.grid(row=2, column=0, ipadx=0, ipady=1, padx=40, sticky=tk.W)

button_frameF = tk.Frame(windowWB)
button_frameF.grid(row=1, column=0, pady=15)

start_buttonF = tk.Button(button_frameF, text="確認", width=6, command=windowWB.destroy)
start_buttonF.grid(row=0, column=0, padx=8)
exit_buttonF = tk.Button(button_frameF, text="關閉", width=6, command=exit_)
exit_buttonF.grid(row=0, column=1, padx=8)

windowWB.mainloop()
"""

"""
import time

Hour = int(time.strftime("%H", time.localtime()))
Min  = int(time.strftime("%M", time.localtime()))
print(type(Hour))
print(Min)

if Hour == 17 and Min == 48:
    while True:
        print("True")
        time.sleep(1)
"""

# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import shutil
import pathlib
import numpy as np
import cv2 as cv

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle

import time
from datetime import datetime

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

from functools import partial
from tenacity import retry, stop_after_delay, wait_fixed, retry_if_exception_type

execute = True

while True:

    # 宣告 "關閉" 按鈕
    def exit_():
        sys.exit()

    def selectExcelPath():
        excelPath_ = askopenfilename()
        excelPath.set(excelPath_)

    if True: # 建立選擇 Excel GUI

        print("請選擇 Excel 路徑\n")

        # 建立使用者 GUI，使其可選擇 Excel 路徑
        windowWB = tk.Tk()
        windowWB.title("Excel 路徑選擇")
        windowWB.geometry('560x120')
        windowWB.resizable(0,0)

        excelPath = StringVar()

        excel_frame2 = tk.Frame(windowWB)
        excel_frame2.grid(row=0, column=0, pady=2)

        excel_label2 = tk.Label(excel_frame2, text="Excel 路徑")
        excel_label2.grid(row=0, column=0, ipadx=10)
        excel_entry2 = tk.Entry(excel_frame2, textvariable=excelPath)
        excel_entry2.grid(row=0, column=1, ipadx=124)
        excel_button2 = tk.Button(excel_frame2, text="選擇", command=selectExcelPath)
        excel_button2.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        button_frameEX = tk.Frame(windowWB)
        button_frameEX.grid(row=1, column=0)

        start_buttonEX = tk.Button(button_frameEX, text="確認", width=6, command=windowWB.destroy)
        start_buttonEX.grid(row=0, column=0, padx=8)
        exit_buttonEX = tk.Button(button_frameEX, text="關閉", width=6, command=exit_)
        exit_buttonEX.grid(row=0, column=1, padx=8)

        windowWB.mainloop()

        print("Excel 路徑為：\n")
        if excelPath.get() != "":
            print(excelPath.get(), "\n")
        print("========================================\n")

    print("Updating formula...  Please Wait...\n")
    print("========================================\n")

    rightNowPath = os.path.join(pathlib.Path().absolute())

    ExcelWBold = openpyxl.load_workbook(excelPath.get())

    resultWSold = ExcelWBold["Results_Overview"]
    for i in range(10, resultWSold.max_row+1):
        print(i)
        resultWSold.delete_rows(10)
    coverageWSold = ExcelWBold["Coverage_Overview"]
    for i in range(10, coverageWSold.max_row+1):
        print(i)
        coverageWSold.delete_rows(10)
    ExcelWBold.save(excelPath.get())