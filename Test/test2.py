"""
a = [2, 4, 1, 5, 3]
for cell in a:
    print(cell)
#print(a)
"""
"""
for i in range(10,1, -1):
    print(i)

from pandas import read_csv
"""
"""
import openpyxl

WB = openpyxl.load_workbook(r"D:\Downloads\Standard Template Datasheet Format.xlsx")
ws = WB["RXL"]
#print(ws["F2"].value)
for row in ws:
    #print(row)
    for cell in row:
        if cell.value == "chart1":
            print(row)
#WB.save(r"D:\Downloads\Standard Template Datasheet Format_new.xlsx")
"""
"""
for i in range (10, 113, 6):
    print(i)
"""


"""
if None == "a":
    print("true")
else:
    print("false")
"""
"""
a = [1, 2, 3, 4, 5, 6]
trueCounter = 0
for i in range(1, 7):
    
    if a[i-1] == i:
        print("true")
        trueCounter += 1
        continue
    else:
        print("false")
        break

if trueCounter == i:
    for i in range(i):
        print(i)
"""
"""
a = []
for i in range(6):
    a.append(i)

print(a)
"""
"""
for i in range(10, 20):
    print(i)
"""
"""
import os
import shutil
#shutil.copyfile(os.path.join("N:/DVT TEST DATA/TB G1.5/TB_G1.5_P1_10_H/6.2_1.1.1_TX_OutputPower_DCA/[MR292190010][+55C][2021-11-12][21_43_22][MTI-PC-1209]/6.2_1.1.1_TX_OutputPower_DCA_PASS", "[TX4][-40.5V@0Hz][DL_71_U-DL_71_V-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide].png"),os.path.join("N:/DVT TEST DATA/TB G1.5 DVT test data/P1/Image/6.2_1.1.1_TX_OutputPower_DCA", "[TX4][-40.5V@0Hz][DL_71_U-DL_71_V-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide]_0010.png"))
#shutil.copyfile(os.path.join(r"N:/DVT TEST DATA/TB G1.5/TB_G1.5_P1_10_H/6.2_1.1.1_TX_OutputPower_DCA/[MR292190010][+55C][2021-11-12][21_43_22][MTI-PC-1209]/6.2_1.1.1_TX_OutputPower_DCA_PASS", r"[TX4][-40.5V@0Hz][DL_71_EF-DL_71_GH-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide].png"),os.path.join(r"N:/DVT TEST DATA/TB G1.5 DVT test data/P1/Image/6.2_1.1.1_TX_OutputPower_DCA", r"[TX4][-40.5V@0Hz][DL_71_EF-DL_71_GH-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide]_0010.png"))

#os.open("N:/DVT TEST DATA/TB G1.5/TB_G1.5_P1_10_H/6.2_1.1.1_TX_OutputPower_DCA/[MR292190010][+55C][2021-11-12][21_43_22][MTI-PC-1209]/6.2_1.1.1_TX_OutputPower_DCA_PASS/[TX4][-40.5V@0Hz][DL_71_EF-DL_71_GH-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide].png")

from PIL import Image                                                                                
img = Image.open("N:/DVT TEST DATA/TB G1.5/TB_G1.5_P1_10_H/6.2_1.1.1_TX_OutputPower_DCA/[MR292190010][+55C][2021-11-12][21_43_22][MTI-PC-1209]/6.2_1.1.1_TX_OutputPower_DCA_PASS/[TX4][-40.5V@0Hz][DL_71_U-DL_71_V-DL_29_A-DL_29_F-DL_26_M5][NR1-TM1.1][Low][DL_71][MXA-LowSide].png")
img.show()
"""
"""
a = ['Passed', 'Passed', 'Passed', 'Passed', 'Skipped', None]
if (None in a) or ("Failed" in a):
    print("YES")
"""
"""
count = 6
for o in range(count):
    print(count)
    print(count+o)
    print(count+o)
"""
"""
import time
print(time.strftime("Start time : %Y/%m/%d %a %H:%M:%S\n", time.localtime()))
print(time.strftime("%H:%M:%S\n", time.localtime()))
print(time.localtime(time.time()))
print(time.localtime())
#time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
#a = time.localtime()
#print(time.localtime() - a)
"""
"""
import time
totalStartTime = time.time()
time.sleep(2)
startTime = time.time()
print((int(time.time() - startTime)))
print((int(time.time() - totalStartTime)))
blockTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(int(time.time() - startTime)))
totalTime = time.strftime("%Y/%m/%d/%H/%M/%S", time.localtime(int(time.time() - totalStartTime)))
print("Block execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(blockTime.split("/")[2])-1, int(blockTime.split("/")[3])-8, int(blockTime.split("/")[4]), int(blockTime.split("/")[5])))
print("Total execute time: %d Day   %d Hr %d Min %d Sec\n" % (int(totalTime.split("/")[2])-1, int(totalTime.split("/")[3])-8, int(totalTime.split("/")[4]), int(totalTime.split("/")[5])))
"""
#"""
#N:\DVT TEST DATA\TB G1.5 DVT test data\P0\Unit_#7J\TB G1.5 P0 DVT test report.xlsx
#N:/DVT TEST DATA/TB G1.5/P0_7J_R
#"""

"""
configArray = []
# 打開 Config
f = open("N:\DVT TEST DATA\ATS Tools\Report Generator\Config.txt", 'r')
fff = f.readline()
print(fff)

for i in range(fff):
    configArray.append(fff)
print(configArray)
"""

"""
if 0==0 and ((0==1) or (0==0)):
    print("True")
"""
"""
import openpyxl

# 存取 Excel 內所有工作表的名稱
ExcelWB = openpyxl.load_workbook("D:\Desktop\Test\DVT test report.xlsx")
ExcelWS = ExcelWB.sheetnames

# 存取 Limit File 內所有工作表的名稱
LimitFileWB = openpyxl.load_workbook("D:\Desktop\Test\LimitFile_7.6_0.1_RX_IBB_SCA.xlsx")
LimitFileWS = LimitFileWB.sheetnames

# 選定要編輯的工作表
MappingWS = LimitFileWB["SheetNameMapping"]

normal = ""
low = ""
high = ""

# 定義路徑欄
for row in MappingWS['A1':'A10']:
    for cell in row:
        if cell.value == None:
            break
        if cell.value[-4:] == "+25C":
            normal = MappingWS.cell(row=cell.row, column=2).value
        if cell.value[-4:] == "-40C":
            low = MappingWS.cell(row=cell.row, column=2).value
        if cell.value[-4:] == "+55C":
            high = MappingWS.cell(row=cell.row, column=2).value

name = ("D:\Desktop\Test\LimitFile_7.6_0.1_RX_IBB_SCA.xlsx").split("\\")[-1][10:-5]

if name in ExcelWS:

    editWS = ExcelWB[name]

    for editWSname in range(10, editWS.max_row+1):
        Allname = []
        if normal != "":
            NormalWS = LimitFileWB[normal]
            for NormalWSname in range(1, NormalWS.max_row+1):
                Allname.append(NormalWS.cell(row=NormalWSname, column=7).value)
        if low != "":
            LowWS = LimitFileWB[low]
            for LowWSname in range(1, LowWS.max_row+1):
                Allname.append(NormalWS.cell(row=LowWSname, column=7).value)
        if high != "":
            HighWS = LimitFileWB[high]
            for HighWSname in range(1, HighWS.max_row+1):
                Allname.append(NormalWS.cell(row=HighWSname, column=7).value)
        if editWS.cell(row=editWSname, column=4).value not in Allname:
            editWS.delete_rows(editWSname)


    if normal != "":
        NormalWS = LimitFileWB[normal]
        for i in range(1, NormalWS.max_row+1):
            TS_No = ("TS_#%.3d" % i)
            Temp = "+25C"
            Test_Name = NormalWS.cell(row=i, column=7).value
            Unit = NormalWS.cell(row=i, column=6).value
            Compare = NormalWS.cell(row=i, column=4).value
            L_LMT = NormalWS.cell(row=i, column=3).value
            H_LMT = NormalWS.cell(row=i, column=2).value

            if Test_Name != editWS.cell(row=i+9, column=4).value:
                editWS.insert_rows(i+9)
                editWS.cell(row=i+9, column=1).value = TS_No
                editWS.cell(row=i+9, column=3).value = Temp
                editWS.cell(row=i+9, column=4).value = Test_Name
                editWS.cell(row=i+9, column=7).value = Unit
                editWS.cell(row=i+9, column=8).value = Compare
                editWS.cell(row=i+9, column=9).value = L_LMT
                editWS.cell(row=i+9, column=10).value = H_LMT
            elif Test_Name == editWS.cell(row=i+9, column=4).value and ((Unit != editWS.cell(row=i+9, column=7).value) or (Compare != editWS.cell(row=i+9, column=8).value) or (L_LMT != editWS.cell(row=i+9, column=9).value) or (H_LMT != editWS.cell(row=i+9, column=10).value)):
                editWS.cell(row=i+9, column=7).value = Unit
                editWS.cell(row=i+9, column=8).value = Compare
                editWS.cell(row=i+9, column=9).value = L_LMT
                editWS.cell(row=i+9, column=10).value = H_LMT

            Counter = i

    if low != "":
        LowWS = LimitFileWB[low]
        for i in range(1, LowWS.max_row+1):
            TS_No = ("TS_#%.3d" % (Counter+i))
            Temp = "-40C"
            Test_Name = LowWS.cell(row=i, column=7).value
            Unit = LowWS.cell(row=i, column=6).value
            Compare = LowWS.cell(row=i, column=4).value
            L_LMT = LowWS.cell(row=i, column=3).value
            H_LMT = LowWS.cell(row=i, column=2).value

            if Test_Name != editWS.cell(row=Counter+i+9, column=4).value:
                editWS.insert_rows(Counter+i+9)
                editWS.cell(row=Counter+i+9, column=1).value = TS_No
                editWS.cell(row=Counter+i+9, column=3).value = Temp
                editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                editWS.cell(row=Counter+i+9, column=7).value = Unit
                editWS.cell(row=Counter+i+9, column=8).value = Compare
                editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                editWS.cell(row=Counter+i+9, column=10).value = H_LMT
            elif Test_Name == editWS.cell(row=Counter+i+9, column=4).value and ((Unit != editWS.cell(row=Counter+i+9, column=7).value) or (Compare != editWS.cell(row=Counter+i+9, column=8).value) or (L_LMT != editWS.cell(row=Counter+i+9, column=9).value) or (H_LMT != editWS.cell(row=Counter+i+9, column=10).value)):
                editWS.cell(row=Counter+i+9, column=7).value = Unit
                editWS.cell(row=Counter+i+9, column=8).value = Compare
                editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                editWS.cell(row=Counter+i+9, column=10).value = H_LMT

            LowCount = i

        Counter = Counter + LowCount

    if high != "":
        HighWS = LimitFileWB[high]
        for i in range(1, HighWS.max_row+1):
            TS_No = ("TS_#%.3d" % (Counter+i))
            Temp = "+55C"
            Test_Name = HighWS.cell(row=i, column=7).value
            Unit = HighWS.cell(row=i, column=6).value
            Compare = HighWS.cell(row=i, column=4).value
            L_LMT = HighWS.cell(row=i, column=3).value
            H_LMT = HighWS.cell(row=i, column=2).value

            if Test_Name != editWS.cell(row=Counter+i+9, column=4).value:
                editWS.insert_rows(Counter+i+9)
                editWS.cell(row=Counter+i+9, column=1).value = TS_No
                editWS.cell(row=Counter+i+9, column=3).value = Temp
                editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                editWS.cell(row=Counter+i+9, column=7).value = Unit
                editWS.cell(row=Counter+i+9, column=8).value = Compare
                editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                editWS.cell(row=Counter+i+9, column=10).value = H_LMT
            elif Test_Name == editWS.cell(row=Counter+i+9, column=4).value and ((Unit != editWS.cell(row=Counter+i+9, column=7).value) or (Compare != editWS.cell(row=Counter+i+9, column=8).value) or (L_LMT != editWS.cell(row=Counter+i+9, column=9).value) or (H_LMT != editWS.cell(row=Counter+i+9, column=10).value)):
                editWS.cell(row=Counter+i+9, column=7).value = Unit
                editWS.cell(row=Counter+i+9, column=8).value = Compare
                editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                editWS.cell(row=Counter+i+9, column=10).value = H_LMT

            HighCount = i

        Counter = Counter + HighCount

else:
    Example = ExcelWB["Example"]
    editWS = ExcelWB.copy_worksheet(Example)
    editWS.title = name

    if normal != "":
        NormalWS = LimitFileWB[normal]
        for i in range(1, NormalWS.max_row+1):
            TS_No = ("TS_#%.3d" % i)
            Temp = "+25C"
            Test_Name = NormalWS.cell(row=i, column=7).value
            Unit = NormalWS.cell(row=i, column=6).value
            Compare = NormalWS.cell(row=i, column=4).value
            L_LMT = NormalWS.cell(row=i, column=3).value
            H_LMT = NormalWS.cell(row=i, column=2).value

            editWS.cell(row=i+9, column=1).value = TS_No
            editWS.cell(row=i+9, column=3).value = Temp
            editWS.cell(row=i+9, column=4).value = Test_Name
            editWS.cell(row=i+9, column=7).value = Unit
            editWS.cell(row=i+9, column=8).value = Compare
            editWS.cell(row=i+9, column=9).value = L_LMT
            editWS.cell(row=i+9, column=10).value = H_LMT

            Counter = i

    if low != "":
        LowWS = LimitFileWB[low]
        for i in range(1, LowWS.max_row+1):
            TS_No = ("TS_#%.3d" % (Counter+i))
            Temp = "-40C"
            Test_Name = LowWS.cell(row=i, column=7).value
            Unit = LowWS.cell(row=i, column=6).value
            Compare = LowWS.cell(row=i, column=4).value
            L_LMT = LowWS.cell(row=i, column=3).value
            H_LMT = LowWS.cell(row=i, column=2).value

            editWS.cell(row=Counter+i+9, column=1).value = TS_No
            editWS.cell(row=Counter+i+9, column=3).value = Temp
            editWS.cell(row=Counter+i+9, column=4).value = Test_Name
            editWS.cell(row=Counter+i+9, column=7).value = Unit
            editWS.cell(row=Counter+i+9, column=8).value = Compare
            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
            editWS.cell(row=Counter+i+9, column=10).value = H_LMT

            LowCount = i

        Counter = Counter + LowCount

    if high != "":
        HighWS = LimitFileWB[high]
        for i in range(1, HighWS.max_row+1):
            TS_No = ("TS_#%.3d" % (Counter+i))
            Temp = "+55C"
            Test_Name = HighWS.cell(row=i, column=7).value
            Unit = HighWS.cell(row=i, column=6).value
            Compare = HighWS.cell(row=i, column=4).value
            L_LMT = HighWS.cell(row=i, column=3).value
            H_LMT = HighWS.cell(row=i, column=2).value

            editWS.cell(row=Counter+i+9, column=1).value = TS_No
            editWS.cell(row=Counter+i+9, column=3).value = Temp
            editWS.cell(row=Counter+i+9, column=4).value = Test_Name
            editWS.cell(row=Counter+i+9, column=7).value = Unit
            editWS.cell(row=Counter+i+9, column=8).value = Compare
            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
            editWS.cell(row=Counter+i+9, column=10).value = H_LMT

            HighCount = i

        Counter = Counter + HighCount

ExcelWB.save("D:\Desktop\Test\DVT test report.xlsx")
"""
"""
import tkinter as tk
import tkinter.ttk as ttk

def combobox_selected(event):
     #print(mycombobox.current(), comboboxText.get())
     labelText.set('' + comboboxText.get())
     
root = tk.Tk()
root.title('my window')
root.geometry('200x150')

comboboxText = tk.StringVar()
mycombobox = ttk.Combobox(root, textvariable=comboboxText, state='readonly')
mycombobox['values'] = ['apple','banana','orange','lemon','tomato']
mycombobox.pack(pady=10)
mycombobox.current(2)

mycombobox.bind('<<ComboboxSelected>>', combobox_selected)

labelText = tk.StringVar()

root.mainloop()

print(labelText.get())
"""
"""
            tempMinus = 0
            totalMinus = editWS.max_row-9
            name = (limitFilePath[executeTime]).split("/")[-1][10:-5]
            if name in ExcelWS:
                editWS = ExcelWB[name]
                print("檢查 Limit File 是否有項目刪減\n")
                for editWSname in range(10, editWS.max_row+1):
                    Allname = []
                    if normal != "":
                        NormalWS = LimitFileWB[normal]
                        for NormalWSname in range(1, NormalWS.max_row+1):
                            Allname.append(NormalWS.cell(row=NormalWSname, column=7).value)
                    if low != "":
                        LowWS = LimitFileWB[low]
                        for LowWSname in range(1, LowWS.max_row+1):
                            Allname.append(NormalWS.cell(row=LowWSname, column=7).value)
                    if high != "":
                        HighWS = LimitFileWB[high]
                        for HighWSname in range(1, HighWS.max_row+1):
                            Allname.append(NormalWS.cell(row=HighWSname, column=7).value)
                    if editWS.cell(row=editWSname, column=4).value not in Allname:
                        editWS.delete_rows(editWSname)
                    # Progress Bar ( 掃描檔案，檢查是否有項目刪減 )
                    #global tempMinus
                    #tempMinus += 1
                    #print('\r' + '%.2f%% |%s%s| %d/%d' % ((float(tempMinus/totalMinus*100)), '█' * int(tempMinus*20/totalMinus), ' ' * (20-int(tempMinus*20/totalMinus)), (editWSname-9), (editWS.max_row-9)), end='')
                print("檢查 Limit File 是否有項目增加\n")
                if normal != "":
                    NormalWS = LimitFileWB[normal]
                    for i in range(1, NormalWS.max_row+1):
                        TS_No = ("TS_#%.3d" % i)
                        Temp = "+25C"
                        Test_Name = NormalWS.cell(row=i, column=7).value
                        Unit = NormalWS.cell(row=i, column=6).value
                        Compare = NormalWS.cell(row=i, column=4).value
                        L_LMT = NormalWS.cell(row=i, column=3).value
                        H_LMT = NormalWS.cell(row=i, column=2).value
                        if Test_Name != editWS.cell(row=i+9, column=4).value:
                            editWS.insert_rows(i+9)
                            editWS.cell(row=i+9, column=1).value = TS_No
                            editWS.cell(row=i+9, column=3).value = Temp
                            editWS.cell(row=i+9, column=4).value = Test_Name
                            editWS.cell(row=i+9, column=7).value = Unit
                            editWS.cell(row=i+9, column=8).value = Compare
                            editWS.cell(row=i+9, column=9).value = L_LMT
                            editWS.cell(row=i+9, column=10).value = H_LMT
                        elif Test_Name == editWS.cell(row=i+9, column=4).value and ((Unit != editWS.cell(row=i+9, column=7).value) or (Compare != editWS.cell(row=i+9, column=8).value) or (L_LMT != editWS.cell(row=i+9, column=9).value) or (H_LMT != editWS.cell(row=i+9, column=10).value)):
                            editWS.cell(row=i+9, column=7).value = Unit
                            editWS.cell(row=i+9, column=8).value = Compare
                            editWS.cell(row=i+9, column=9).value = L_LMT
                            editWS.cell(row=i+9, column=10).value = H_LMT
                        Counter = i
                if low != "":
                    LowWS = LimitFileWB[low]
                    for i in range(1, LowWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+i))
                        Temp = "-40C"
                        Test_Name = LowWS.cell(row=i, column=7).value
                        Unit = LowWS.cell(row=i, column=6).value
                        Compare = LowWS.cell(row=i, column=4).value
                        L_LMT = LowWS.cell(row=i, column=3).value
                        H_LMT = LowWS.cell(row=i, column=2).value
                        if Test_Name != editWS.cell(row=Counter+i+9, column=4).value:
                            editWS.insert_rows(Counter+i+9)
                            editWS.cell(row=Counter+i+9, column=1).value = TS_No
                            editWS.cell(row=Counter+i+9, column=3).value = Temp
                            editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                            editWS.cell(row=Counter+i+9, column=7).value = Unit
                            editWS.cell(row=Counter+i+9, column=8).value = Compare
                            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                            editWS.cell(row=Counter+i+9, column=10).value = H_LMT
                        elif Test_Name == editWS.cell(row=Counter+i+9, column=4).value and ((Unit != editWS.cell(row=Counter+i+9, column=7).value) or (Compare != editWS.cell(row=Counter+i+9, column=8).value) or (L_LMT != editWS.cell(row=Counter+i+9, column=9).value) or (H_LMT != editWS.cell(row=Counter+i+9, column=10).value)):
                            editWS.cell(row=Counter+i+9, column=7).value = Unit
                            editWS.cell(row=Counter+i+9, column=8).value = Compare
                            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                            editWS.cell(row=Counter+i+9, column=10).value = H_LMT
                        LowCount = i
                    Counter = Counter + LowCount
                if high != "":
                    HighWS = LimitFileWB[high]
                    for i in range(1, HighWS.max_row+1):
                        TS_No = ("TS_#%.3d" % (Counter+i))
                        Temp = "+55C"
                        Test_Name = HighWS.cell(row=i, column=7).value
                        Unit = HighWS.cell(row=i, column=6).value
                        Compare = HighWS.cell(row=i, column=4).value
                        L_LMT = HighWS.cell(row=i, column=3).value
                        H_LMT = HighWS.cell(row=i, column=2).value
                        if Test_Name != editWS.cell(row=Counter+i+9, column=4).value:
                            editWS.insert_rows(Counter+i+9)
                            editWS.cell(row=Counter+i+9, column=1).value = TS_No
                            editWS.cell(row=Counter+i+9, column=3).value = Temp
                            editWS.cell(row=Counter+i+9, column=4).value = Test_Name
                            editWS.cell(row=Counter+i+9, column=7).value = Unit
                            editWS.cell(row=Counter+i+9, column=8).value = Compare
                            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                            editWS.cell(row=Counter+i+9, column=10).value = H_LMT
                        elif Test_Name == editWS.cell(row=Counter+i+9, column=4).value and ((Unit != editWS.cell(row=Counter+i+9, column=7).value) or (Compare != editWS.cell(row=Counter+i+9, column=8).value) or (L_LMT != editWS.cell(row=Counter+i+9, column=9).value) or (H_LMT != editWS.cell(row=Counter+i+9, column=10).value)):
                            editWS.cell(row=Counter+i+9, column=7).value = Unit
                            editWS.cell(row=Counter+i+9, column=8).value = Compare
                            editWS.cell(row=Counter+i+9, column=9).value = L_LMT
                            editWS.cell(row=Counter+i+9, column=10).value = H_LMT
                        HighCount = i
                    Counter = Counter + HighCount
"""


import openpyxl
TestDataWB = openpyxl.load_workbook(r"C:\Users\6476\Desktop\[3LMTC90007J][+25C][2022-01-25][16_10_48][MTI-PC-1210]\(3LMTC90007J)(+25C)(2022-01-25)(16_10_48)(MTI-PC-1210)_TestReport.xlsx")
if "6.2_1.1.1_TX_OutputPower_DCA" not in TestDataWB.sheetnames:
    print("error")
    print(TestDataWB.sheetnames)
#TestDataWS = TestDataWB["6.2_1.1.1_TX_OutputPower_DCA"]
#print(TestDataWS)