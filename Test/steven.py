import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from os import listdir,getcwd
from os.path import isfile, join, basename, splitext, dirname, abspath
from openpyxl import load_workbook
from pandas import read_csv

root = tk.Tk()
root.withdraw()
#file_path = filedialog.askdirectory(parent=root,initialdir='os.path.dirname(os.path.abspath(__file__))')
file_path = filedialog.askdirectory(parent=root,initialdir=dirname(abspath(__file__)))
print(file_path)
fileExt = r".csv"
files = listdir(file_path)

# 以迴圈處理
for f in files:
    # 產生檔案的絕對路徑
    fullpath = join(file_path, f)
    # 判斷 fullpath 是檔案還是目錄
    if isfile(fullpath):
        print("Source file：", f)
        if f.lower().endswith(fileExt):
            csv = read_csv(fullpath, encoding='utf-8')
            csv.to_excel(join(getcwd(),'csv to excel.xlsx'), sheet_name='Template Data')
            wb_csv = load_workbook(join(getcwd(),'csv to excel.xlsx'))
            sheet_csv = wb_csv['Template Data']
            wb = load_workbook(join(getcwd(),'Standard Template Datasheet Format.xlsx'))
            sheet = wb['Template Data']
            for row in sheet_csv:
                for cell in row:
                    sheet[cell.coordinate].value = cell.value
            base_name = basename(f)
            file_name = splitext(base_name)[0]
            #wb.save(join(getcwd(),'Transform',file_name)+'_update.xlsx')
            print("Transform file：", join(getcwd(),'Transform',file_name)+'_update.xlsx')
messagebox.showinfo('messagebox', 'Transform ok')