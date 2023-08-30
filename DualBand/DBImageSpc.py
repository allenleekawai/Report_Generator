# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import sys
import shutil
import openpyxl
from openpyxl.styles import Font
import time
from datetime import datetime

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

from functools import partial
from tenacity import retry, stop_after_delay, wait_fixed, retry_if_exception_type

execute = True

while True:

    # ================================================================================ #
    #                                                                                  #
    #                             宣告 Function 及建立 GUI                              #
    #                                                                                  #
    # ================================================================================ #

    if True:

        # 宣告 "選擇Excel" Function
        def selectExcelPath():
            excelPath_ = askopenfilename()
            excelPath.set(excelPath_)

        # 宣告 "選擇資料夾" Function
        def selectFolderPath1():
            folderPath_1 = askdirectory()
            folderPath1.set(folderPath_1)

        # 宣告 "關閉" 按鈕
        def exit_():
            sys.exit()

        class CommunicationError(Exception):
            pass

        retry_on_communication_error = partial(retry,
            stop=stop_after_delay(20),  # max. 20 seconds wait.
            wait=wait_fixed(2),  # wait 2 seconds
            retry=retry_if_exception_type(CommunicationError),
            reraise=True
        )()

        # ================================================================================ #
        #                                                                                  #
        #                                     建立 GUI                                     #
        #                                                                                  #
        # ================================================================================ #

        print("Please select Excel path\n")

        # 建立使用者 GUI，使其可選擇 Excel 路徑
        windowWB = tk.Tk()
        windowWB.title("Excel select")
        windowWB.geometry('550x89')
        windowWB.resizable(0,0)

        excelPath  = StringVar()

        excel_frame = tk.Frame(windowWB)
        excel_frame.grid(row=0, column=0, pady=2)

        excel_label = tk.Label(excel_frame, text="Excel path")
        excel_label.grid(row=0, column=0, ipadx=10)
        excel_entry = tk.Entry(excel_frame, textvariable=excelPath)
        excel_entry.grid(row=0, column=1, ipadx=124)
        excel_button = tk.Button(excel_frame, text="select", command=selectExcelPath)
        excel_button.grid(row=0, column=2, padx=7, pady=4)

        button_frameWB = tk.Frame(windowWB)
        button_frameWB.grid(row=1, column=0)

        start_buttonWB = tk.Button(button_frameWB, text="OK", width=6, command=windowWB.destroy)
        start_buttonWB.grid(row=0, column=0, padx=8)
        exit_buttonWB = tk.Button(button_frameWB, text="Close", width=6, command=exit_)
        exit_buttonWB.grid(row=0, column=1, padx=8)

        windowWB.mainloop()

        print("Path is: ", excelPath.get(), "\n")
        print("========================================\n")

        # ================================================================================ #
        # ================================================================================ #

        print("Please select Image path\n")

        # 建立使用者 GUI，使其可選擇要導入 Image 的資料夾
        windowImg = tk.Tk()
        windowImg.title("Image Path select")
        windowImg.geometry('570x89')
        windowImg.resizable(0,0)

        folderPath1 = StringVar()

        # ==================== #

        image_frame1 = tk.Frame(windowImg)
        image_frame1.grid(row=0, column=0, pady=2)

        image_label1 = tk.Label(image_frame1, text="Image path")
        image_label1.grid(row=0, column=0, ipadx=10)
        image_entry1 = tk.Entry(image_frame1, textvariable=folderPath1)
        image_entry1.grid(row=0, column=1, ipadx=124)
        image_button1 = tk.Button(image_frame1, text="Select", command=selectFolderPath1)
        image_button1.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #


        button_frameImg = tk.Frame(windowImg)
        button_frameImg.grid(row=9, column=0)

        start_buttonImg = tk.Button(button_frameImg, text="OK", width=6, command=windowImg.destroy)
        start_buttonImg.grid(row=0, column=0, padx=8, pady=10)
        exit_buttonImg = tk.Button(button_frameImg, text="Close", width=6, command=exit_)
        exit_buttonImg.grid(row=0, column=1, padx=8, pady=10)

        windowWB.mainloop()
        folderPath = []

        print("Path is: \n")
        if folderPath1.get() != "":
            print(folderPath1.get(), "\n")
            folderPath.append(folderPath1.get())
        print("========================================\n")

    # ================================================================================ #
    #                                                                                  #
    #                                   建立前置作業                                    #
    #                                                                                  #
    # ================================================================================ #

    if True:
        # 存取 Excel 內所有工作表的名稱
        wb = openpyxl.load_workbook(excelPath.get())
        
        ws = wb.sheetnames
        ws.remove("Results_Overview")
        ws.remove("Coverage_Overview")
        ws.remove("Example")
        ws.remove("1.0_0.1.1_Power_Consump_SCA")
        ws.remove("1.1_0.1.1_RX_RSSI_SCA")
        
        #"""
        # 指定 Image 資料夾
        newImagePath = os.path.dirname(excelPath.get()) + "/Image"

        itemName = ""
        for worksheet in ws:
            # 選定要編輯的工作表
            editWS = wb[worksheet]
            itemName = worksheet

            newImageFolder = ""
            # 指定測項子 Image 資料夾
            dir = os.listdir(newImagePath)
            print(dir)
            print(itemName)
            for j in dir:
                if j == itemName:
                    newImageFolder = newImagePath + "/" + j

            picDir = os.listdir(newImageFolder)
            print(picDir)

            # 定義路徑欄
            for row in editWS['A9':'AZ9']:
                for cell in row:
                    unitCoordinate = 2
                    if cell.value == "Temp.":
                        tempCoordinate = cell.column
                    if cell.value == "Test_Name":
                        nameCoordinate = cell.column
                    if cell.value == "Image":
                        imageCoordinate = cell.column
                    if cell.value == "Session":
                        sessionCoordinate = cell.column
                    if cell.value == "Start_Time":
                        timeCoordinate = cell.column
                    if cell.value == "Test Plot":
                        picCoordinate = cell.column

            # 宣告最後用來放入 Pic 超連結路徑的 Array
            picName     = ["Name"] * (editWS.max_row + 1 - 10)
            testName    = ["TestName"] * (editWS.max_row + 1 - 10)
            newTestName = ["NewName"] * (editWS.max_row + 1 - 10)

            # ================================================================================ #
            #                                                                                  #
            #                      掃描資料夾，建立 Excel Test Plot 超連結                       #
            #                                                                                  #
            # ================================================================================ #

            print("Searching file, Creating hyperlink...\n")

            temp1 = 0
            total1 = editWS.max_row-9

            finalFolderPath = ""

            @retry_on_communication_error
            def HyperLinkBuilding(i):

                global finalFolderPath
                finalFolderPath = ""

                # 儲存照片名稱欄位
                picOrg = (editWS.cell(row=i, column=picCoordinate).value)

                # 去掉 Test Name 前面的項目名，只留下後段測試名
                testName[i-10] = (editWS.cell(row=i, column=nameCoordinate).value)
                nameSplit = testName[i-10].split("][")
                newTestName[i-10] = "[" + (nameSplit[0].split("["))[1]
                for nameSplitCount in range(1, len(nameSplit)):
                    newTestName[i-10] = newTestName[i-10] + "][" + nameSplit[nameSplitCount]

                for k in picDir:
                    #print(k)
                    #print(newTestName[i-10])
                    #print(picOrg)
                    if k[-4:] == ".png": 
                        if (picOrg!=None and picOrg!="") and (newTestName[i-10]!=None and newTestName[i-10]!=""):
                            if (newTestName[i-10] in k) or (picOrg in k):
                                picName[i-10] = k

                # Progress Bar One ( 掃描檔案，建立超連結 )
                global temp1
                temp1 += 1
                print('\r' + '%.2f%% |%s%s| %d/%d' % ((float(temp1/total1*100)), '█' * int(temp1*20/total1), ' ' * (20-int(temp1*20/total1)), (i-9), (editWS.max_row-9)), end='\n')

            for i in range(10, editWS.max_row+1):
                HyperLinkBuilding(i)

            # ================================================================================ #
            #                                                                                  #
            #                      複製照片至 Excel 路徑底下的 Image 資料夾                      #
            #                                                                                  #
            # ================================================================================ #

            print("\n\nCopying image...\n")

            temp2 = 0
            total2 = editWS.max_row-9

            @retry_on_communication_error
            def ImageCopy(i):
                unit = (editWS.cell(row=i, column=unitCoordinate).value)
                temp = (editWS.cell(row=i, column=tempCoordinate).value)

                if picName[i-10] != "Name" and ((editWS.cell(row=i, column=imageCoordinate).value == "") or (editWS.cell(row=i, column=imageCoordinate).value == None)):
                    editWS.cell(row=i, column=imageCoordinate).value = ('=HYPERLINK("%s", "Link")' % picName[i-10])
                    editWS.cell(row=i, column=imageCoordinate).font = Font(underline="single", color='00B050')

                # Progress Bar Two ( 照片複製 )
                global temp2
                temp2 += 1
                print('\r' + '%.2f%% |%s%s| %d/%d' % ((float(temp2/total2*100)), '█' * int(temp2*20/total2), ' ' * (20-int(temp2*20/total2)), (i-9), (editWS.max_row-9)), end='')

            for i in range(10, editWS.max_row+1):
                ImageCopy(i)

            print("\n\nSaving...  Please Wait...\n")
            print("========================================\n")
            wb.save(excelPath.get())

    # ================================================================================ #
    #                                                                                  #
    #                              檔案儲存 ( Excel, txt )                              #
    #                                                                                  #
    # ================================================================================ #

    print("\n\nSaving...  Please Wait...\n")
    print("========================================\n")
    wb.save(excelPath.get())
    print("Success!\n")
    time.sleep(2)

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    print("========================================\n")
    print("Re-run Y/N ?")
    executeYN = input()
    print()
    print("========================================\n")

    if executeYN == "Y" or executeYN == "y":
        execute = True
    elif executeYN == "N" or executeYN == "n":
        execute = False
    if execute == False:
        break
#"""
