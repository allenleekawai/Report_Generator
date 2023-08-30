# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import shutil
import openpyxl
from openpyxl.styles import Font
import time
from datetime import datetime

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename

from functools import partial
from tenacity import retry, stop_after_delay, wait_fixed, retry_if_exception_type

execute = True

while True:

    # ================================================================================ #
    #                                                                                  #
    #                             宣告 Function 及建立 GUI                              #
    #                                                                                  #
    # ================================================================================ #

    if True:

        # 宣告 "選擇Excel" Function
        def selectExcelPath():
            excelPath_ = askopenfilename()
            excelPath.set(excelPath_)

        # 宣告 "選擇資料夾" Function
        def selectFolderPath1():
            folderPath_1 = askdirectory()
            folderPath1.set(folderPath_1)

        def selectFolderPath2():
            folderPath_2 = askdirectory()
            folderPath2.set(folderPath_2)

        def selectFolderPath3():
            folderPath_3 = askdirectory()
            folderPath3.set(folderPath_3)

        def selectFolderPath4():
            folderPath_4 = askdirectory()
            folderPath4.set(folderPath_4)

        def selectFolderPath5():
            folderPath_5 = askdirectory()
            folderPath5.set(folderPath_5)

        def selectFolderPath6():
            folderPath_6 = askdirectory()
            folderPath6.set(folderPath_6)

        def selectFolderPath7():
            folderPath_7 = askdirectory()
            folderPath7.set(folderPath_7)

        def selectFolderPath8():
            folderPath_8 = askdirectory()
            folderPath8.set(folderPath_8)

        def selectFolderPath9():
            folderPath_9 = askdirectory()
            folderPath9.set(folderPath_9)

        # 宣告 "下拉式選單" Function
        def combobox_selected(event):
            labelWSText.set(comboboxWSText.get())

        # 宣告 "Error Log 編寫" Function
        def error(text):
            f.write(text)
            f.write("\n")

        # 宣告 "關閉" 按鈕
        def exit_():
            exit()

        class CommunicationError(Exception):
            pass

        retry_on_communication_error = partial(retry,
            stop=stop_after_delay(20),  # max. 20 seconds wait.
            wait=wait_fixed(2),  # wait 2 seconds
            retry=retry_if_exception_type(CommunicationError),
            reraise=True
        )()

        # ================================================================================ #
        #                                                                                  #
        #                                     建立 GUI                                     #
        #                                                                                  #
        # ================================================================================ #

        print("請選擇 Excel 路徑\n")

        # 建立使用者 GUI，使其可選擇 Excel 路徑
        windowWB = tk.Tk()
        windowWB.title("Excel選擇")
        windowWB.geometry('550x89')
        windowWB.resizable(0,0)

        excelPath  = StringVar()

        excel_frame = tk.Frame(windowWB)
        excel_frame.grid(row=0, column=0, pady=2)

        excel_label = tk.Label(excel_frame, text="Excel路徑")
        excel_label.grid(row=0, column=0, ipadx=10)
        excel_entry = tk.Entry(excel_frame, textvariable=excelPath)
        excel_entry.grid(row=0, column=1, ipadx=124)
        excel_button = tk.Button(excel_frame, text="選擇", command=selectExcelPath)
        excel_button.grid(row=0, column=2, padx=7, pady=4)

        button_frameWB = tk.Frame(windowWB)
        button_frameWB.grid(row=1, column=0)

        start_buttonWB = tk.Button(button_frameWB, text="確認", width=6, command=windowWB.destroy)
        start_buttonWB.grid(row=0, column=0, padx=8)
        exit_buttonWB = tk.Button(button_frameWB, text="關閉", width=6, command=exit_)
        exit_buttonWB.grid(row=0, column=1, padx=8)

        windowWB.mainloop()

        print("路徑為：", excelPath.get(), "\n")
        print("========================================\n")

        # 存取 Excel 內所有工作表的名稱
        wb = openpyxl.load_workbook(excelPath.get())
        ws = wb.sheetnames

        # ================================================================================ #
        # ================================================================================ #

        print("請選擇要導入 Image 的工作表\n")

        # 建立使用者 GUI，使其可選擇要導入 Image 的工作表
        windowWS = tk.Tk()
        windowWS.title("工作表選擇")
        windowWS.geometry('440x280')
        windowWS.resizable(0,0)

        comboboxWSText = StringVar()
        labelWSText = StringVar()

        labelTop = tk.Label(windowWS, text = "選擇要導入 Image 的工作表")
        labelTop.grid(column=0, row=0, pady=10)

        comboboxWS = ttk.Combobox(windowWS, values=ws, width=50, state='readonly', textvariable=comboboxWSText)
        comboboxWS.grid(column=0, row=1, padx=15)
        #comboboxWS.current(0)
        comboboxWS.bind('<<ComboboxSelected>>', combobox_selected)

        labelChoose = tk.Label(windowWS, text="選擇的工作表是：")
        labelChoose.grid(column=0, row=2, pady=16)

        labelWS = tk.Label(windowWS, textvariable=labelWSText, font=('Arial', 18))
        labelWS.grid(column=0, row=3)

        button_frameWS = tk.Frame(windowWS)
        button_frameWS.grid(column=0, row=4, pady=45)

        start_buttonWS = tk.Button(button_frameWS, text="確認", width=6, command=windowWS.destroy)
        start_buttonWS.grid(column=0, row=0, padx=8)
        exit_buttonWS = tk.Button(button_frameWS, text="關閉", width=6, command=exit_)
        exit_buttonWS.grid(column=1, row=0, padx=8)

        windowWS.mainloop()

        print("選擇了：", comboboxWSText.get(), "\n")
        print("========================================\n")

        # ================================================================================ #
        # ================================================================================ #

        print("請選擇導入 Image 的路徑\n")

        # 建立使用者 GUI，使其可選擇要導入 Image 的資料夾
        windowImg = tk.Tk()
        windowImg.title("Image 路徑選擇")
        windowImg.geometry('570x430')
        windowImg.resizable(0,0)

        folderPath1 = StringVar()
        folderPath2 = StringVar()
        folderPath3 = StringVar()
        folderPath4 = StringVar()
        folderPath5 = StringVar()
        folderPath6 = StringVar()
        folderPath7 = StringVar()
        folderPath8 = StringVar()
        folderPath9 = StringVar()

        # ==================== #

        image_frame1 = tk.Frame(windowImg)
        image_frame1.grid(row=0, column=0, pady=2)

        image_label1 = tk.Label(image_frame1, text="Image 路徑1")
        image_label1.grid(row=0, column=0, ipadx=10)
        image_entry1 = tk.Entry(image_frame1, textvariable=folderPath1)
        image_entry1.grid(row=0, column=1, ipadx=124)
        image_button1 = tk.Button(image_frame1, text="選擇", command=selectFolderPath1)
        image_button1.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame2 = tk.Frame(windowImg)
        image_frame2.grid(row=1, column=0, pady=2)

        image_label2 = tk.Label(image_frame2, text="Image 路徑2")
        image_label2.grid(row=0, column=0, ipadx=10)
        image_entry2 = tk.Entry(image_frame2, textvariable=folderPath2)
        image_entry2.grid(row=0, column=1, ipadx=124)
        image_button2 = tk.Button(image_frame2, text="選擇", command=selectFolderPath2)
        image_button2.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame3 = tk.Frame(windowImg)
        image_frame3.grid(row=2, column=0, pady=2)

        image_label3 = tk.Label(image_frame3, text="Image 路徑3")
        image_label3.grid(row=0, column=0, ipadx=10)
        image_entry3 = tk.Entry(image_frame3, textvariable=folderPath3)
        image_entry3.grid(row=0, column=1, ipadx=124)
        image_button3 = tk.Button(image_frame3, text="選擇", command=selectFolderPath3)
        image_button3.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame4 = tk.Frame(windowImg)
        image_frame4.grid(row=3, column=0, pady=2)

        image_label4 = tk.Label(image_frame4, text="Image 路徑4")
        image_label4.grid(row=0, column=0, ipadx=10)
        image_entry4 = tk.Entry(image_frame4, textvariable=folderPath4)
        image_entry4.grid(row=0, column=1, ipadx=124)
        image_button4 = tk.Button(image_frame4, text="選擇", command=selectFolderPath4)
        image_button4.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame5 = tk.Frame(windowImg)
        image_frame5.grid(row=4, column=0, pady=2)

        image_label5 = tk.Label(image_frame5, text="Image 路徑5")
        image_label5.grid(row=0, column=0, ipadx=10)
        image_entry5 = tk.Entry(image_frame5, textvariable=folderPath5)
        image_entry5.grid(row=0, column=1, ipadx=124)
        image_button5 = tk.Button(image_frame5, text="選擇", command=selectFolderPath5)
        image_button5.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame6 = tk.Frame(windowImg)
        image_frame6.grid(row=5, column=0, pady=2)

        image_label6 = tk.Label(image_frame6, text="Image 路徑6")
        image_label6.grid(row=0, column=0, ipadx=10)
        image_entry6 = tk.Entry(image_frame6, textvariable=folderPath6)
        image_entry6.grid(row=0, column=1, ipadx=124)
        image_button6 = tk.Button(image_frame6, text="選擇", command=selectFolderPath6)
        image_button6.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame7 = tk.Frame(windowImg)
        image_frame7.grid(row=6, column=0, pady=2)

        image_label7 = tk.Label(image_frame7, text="Image 路徑7")
        image_label7.grid(row=0, column=0, ipadx=10)
        image_entry7 = tk.Entry(image_frame7, textvariable=folderPath7)
        image_entry7.grid(row=0, column=1, ipadx=124)
        image_button7 = tk.Button(image_frame7, text="選擇", command=selectFolderPath7)
        image_button7.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame8 = tk.Frame(windowImg)
        image_frame8.grid(row=7, column=0, pady=2)

        image_label8 = tk.Label(image_frame8, text="Image 路徑8")
        image_label8.grid(row=0, column=0, ipadx=10)
        image_entry8 = tk.Entry(image_frame8, textvariable=folderPath8)
        image_entry8.grid(row=0, column=1, ipadx=124)
        image_button8 = tk.Button(image_frame8, text="選擇", command=selectFolderPath8)
        image_button8.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #

        image_frame9 = tk.Frame(windowImg)
        image_frame9.grid(row=8, column=0, pady=2)

        image_label9 = tk.Label(image_frame9, text="Image 路徑9")
        image_label9.grid(row=0, column=0, ipadx=10)
        image_entry9 = tk.Entry(image_frame9, textvariable=folderPath9)
        image_entry9.grid(row=0, column=1, ipadx=124)
        image_button9 = tk.Button(image_frame9, text="選擇", command=selectFolderPath9)
        image_button9.grid(row=0, column=2, padx=7, pady=4)

        # ==================== #


        button_frameImg = tk.Frame(windowImg)
        button_frameImg.grid(row=9, column=0)

        start_buttonImg = tk.Button(button_frameImg, text="確認", width=6, command=windowImg.destroy)
        start_buttonImg.grid(row=0, column=0, padx=8, pady=10)
        exit_buttonImg = tk.Button(button_frameImg, text="關閉", width=6, command=exit_)
        exit_buttonImg.grid(row=0, column=1, padx=8, pady=10)

        windowWB.mainloop()
        folderPath = []

        print("路徑為：\n")
        if folderPath1.get() != "":
            print(folderPath1.get(), "\n")
            folderPath.append(folderPath1.get())
        if folderPath2.get() != "":
            print(folderPath2.get(), "\n")
            folderPath.append(folderPath2.get())
        if folderPath3.get() != "":
            print(folderPath3.get(), "\n")
            folderPath.append(folderPath3.get())
        if folderPath4.get() != "":
            print(folderPath4.get(), "\n")
            folderPath.append(folderPath4.get())
        if folderPath5.get() != "":
            print(folderPath5.get(), "\n")
            folderPath.append(folderPath5.get())
        if folderPath6.get() != "":
            print(folderPath6.get(), "\n")
            folderPath.append(folderPath6.get())
        if folderPath7.get() != "":
            print(folderPath7.get(), "\n")
            folderPath.append(folderPath7.get())
        if folderPath8.get() != "":
            print(folderPath8.get(), "\n")
            folderPath.append(folderPath8.get())
        if folderPath9.get() != "":
            print(folderPath9.get(), "\n")
            folderPath.append(folderPath9.get())
        print("========================================\n")

    # ================================================================================ #
    #                                                                                  #
    #                                   建立前置作業                                    #
    #                                                                                  #
    # ================================================================================ #

    if True:
        # 查看 Image 資料夾是否存在，若無，建立 Image 資料夾
        newImagePath = os.path.dirname(excelPath.get())
        if not(os.path.exists(newImagePath + "/Image")):
            os.makedirs(newImagePath + "/Image")

        # 查看測項子 Image 資料夾是否存在，若無，建立測項子 Image 資料夾
        newImageFolder = newImagePath + "/Image/" + comboboxWSText.get()
        if not(os.path.exists(newImageFolder)):
            os.makedirs(newImageFolder)

        # 創建 Error Log
        txtPath = newImagePath + "/" + (excelPath.get()).split("/")[-1][:-5] + '_ErrorLog.txt'
        if os.path.isfile(txtPath):
            os.remove(txtPath)
        f = open(txtPath, 'w')
        f.write("在資料夾中找到照片\n但因伺服器網路異常\n只有 HyperLink 被建立\n照片並未複製\n\nWrong File:\n\n")

        # 選定要編輯的工作表
        editWS = wb[comboboxWSText.get()]

        # 定義路徑欄
        for row in editWS['A9':'AZ9']:
            for cell in row:
                unitCoordinate = 2
                if cell.value == "Temp.":
                    tempCoordinate = cell.column
                if cell.value == "Test_Name":
                    nameCoordinate = cell.column
                if cell.value == "Image":
                    imageCoordinate = cell.column
                if cell.value == "Session":
                    sessionCoordinate = cell.column
                if cell.value == "Start_Time":
                    timeCoordinate = cell.column
                if cell.value == "Test Plot":
                    picCoordinate = cell.column

        # 宣告最後用來放入 Pic 超連結路徑的 Array
        picPath = ["Path"] * (editWS.max_row + 1 - 10)
        picName = ["Name"] * (editWS.max_row + 1 - 10)
        # 宣告最後用來放入 Session 的 Array
        sessionPath = ["Path"] * (editWS.max_row + 1 - 10)

        # 將所有符合工作表名稱的資料夾放入 folder[]
        folder = []

        for i in range(len(folderPath)):
            dir = os.listdir(folderPath[i])
            for j in range(len(dir)):
                if comboboxWSText.get() == dir[j]:  folder.append(folderPath[i] + "/" + dir[j])

    # ================================================================================ #
    #                                                                                  #
    #                      掃描資料夾，建立 Excel Test Plot 超連結                       #
    #                                                                                  #
    # ================================================================================ #

    print("掃描檔案，建立超連結中...\n")

    temp1 = 0
    total1 = editWS.max_row-9

    finalFolderPath = ""

    @retry_on_communication_error
    def HyperLinkBuilding(i):

        global finalFolderPath
        finalFolderPath = ""
        
        # 儲存照片名稱欄位
        picOrg = (editWS.cell(row=i, column=picCoordinate).value)

        if picOrg != None:
            if editWS.cell(row=i, column=imageCoordinate).value == None:

                # 儲存 Unit 欄位
                unit = (editWS.cell(row=i, column=unitCoordinate).value)
                # 儲存溫度欄位
                temp = (editWS.cell(row=i, column=tempCoordinate).value)
                # 儲存開始時間，並將 月日時分秒 按公式轉換成秒
                startTime = (editWS.cell(row=i, column=timeCoordinate).value)
                startTimeSec = time.mktime((datetime.strptime(startTime, "%Y-%m-%d_%H:%M:%S")).timetuple())

                if "RX_IBB_SCA" in (comboboxWSText.get())\
                or "RX_IM_SCA" in (comboboxWSText.get())\
                or "RX_COL_SCA" in (comboboxWSText.get())\
                or "RX_ACS_SCA" in (comboboxWSText.get())\
                or "RX_NBB_SCA" in (comboboxWSText.get())\
                or "RX_NF_SCA" in (comboboxWSText.get())\
                or "RX_OOB_SCA" in (comboboxWSText.get())\
                or "TX_ACLR_DCA" in (comboboxWSText.get()):
                    pic = picOrg + ".png"
                    for j in range(len(folder)):
                        compare = 10000000000.0
                        # 讀取所有符合工作表名稱資料夾的子資料夾
                        dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                        for k in range(len(dir)):
                            # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                            if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                # 掃描最接近測項時間的資料夾
                                if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                    compare = startTimeSec - folderTimeSec
                                    finalFolderPath = folder[j] + "/" + dir[k]

                elif len(picOrg.split("][")) == 10:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[4] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[4]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[5] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[5]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[6] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[6]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[7] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[7]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[8] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[8]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 9:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[4] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[4]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[5] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[5]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[6] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[6]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[7] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[7]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 8:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[4] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[4]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[5] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[5]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[6] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[6]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 7:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[4] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[4]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[5] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[5]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 6:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[4] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[4]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 5:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[3] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[3]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 4:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[2] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[2]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                elif len(picOrg.split("][")) == 3:
                    pic = picOrg + ".png"
                    # 比對照片名稱與 Test Name 是否相符，若不相符，可能為人工貼表時產生錯誤
                    if (editWS.cell(row=i, column=picCoordinate).value).split("][")[0][-3:] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[0][-3:]\
                        and (editWS.cell(row=i, column=picCoordinate).value).split("][")[1] == (editWS.cell(row=i, column=nameCoordinate).value).split("][")[1]:
                            for j in range(len(folder)):
                                compare = 10000000000.0
                                # 讀取所有符合工作表名稱資料夾的子資料夾
                                dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                                for k in range(len(dir)):
                                    # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                                    if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                        folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                        folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                        # 掃描最接近測項時間的資料夾
                                        if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                            compare = startTimeSec - folderTimeSec
                                            finalFolderPath = folder[j] + "/" + dir[k]
                    else:
                        print("第%d行，照片名稱與 Test Name 不相符，可能為人工貼表時產生錯誤" % i)

                else:
                    pic = picOrg + ".png"
                    for j in range(len(folder)):
                        compare = 10000000000.0
                        # 讀取所有符合工作表名稱資料夾的子資料夾
                        dir = ([ name for name in os.listdir(folder[j]) if os.path.isdir(os.path.join(folder[j], name)) ])
                        for k in range(len(dir)):
                            # 掃描子資料夾的 Unit, Temp 與測項是否吻合
                            if unit == (dir[k][1:-1].split("]["))[0] and temp == (dir[k][1:-1].split("]["))[1]:
                                folderTimeStr = (dir[k].split("][")[2]) + "_" + (dir[k].split("][")[3])
                                folderTimeSec = time.mktime((datetime.strptime(folderTimeStr, "%Y-%m-%d_%H_%M_%S")).timetuple())
                                # 掃描最接近測項時間的資料夾
                                if startTimeSec > folderTimeSec and (startTimeSec-folderTimeSec) < compare:
                                    compare = startTimeSec - folderTimeSec
                                    finalFolderPath = folder[j] + "/" + dir[k]

                # 編寫 Session 路徑
                sessionPath[i - 10] = finalFolderPath.split("/")[-1]

                #print(finalFolderPath)
                finalDir = ([ name for name in os.listdir(finalFolderPath) if os.path.isdir(os.path.join(finalFolderPath, name)) ])

                for x in range(len(finalDir)):
                    # 找尋放照片的資料夾
                    if finalDir[x][-5:] == "_PASS" or finalDir[x][-5:] == "_FAIL":
                        if comboboxWSText.get() == finalDir[x][:-5]:
                            picDir = os.listdir(finalFolderPath + "/" + finalDir[x])
                            for fichier in picDir[:]:
                                if not(fichier.endswith(".png")):
                                    picDir.remove(fichier)
                            for y in range(len(picDir)):
                                # 若檔案名稱與 Excel 內照片名稱欄的名字一樣，則放入 picPath array
                                if pic == picDir[y]:
                                    picPath[i - 10] = (finalFolderPath + "/" + finalDir[x])
                                    picName[i - 10] = picDir[y]
                    else:
                        if comboboxWSText.get() == finalDir[x]:
                            picDir = os.listdir(finalFolderPath + "/" + finalDir[x])
                            for fichier in picDir[:]:
                                if not(fichier.endswith(".png")):
                                    picDir.remove(fichier)
                            for y in range(len(picDir)):
                                # 若檔案名稱與 Excel 內照片名稱欄的名字一樣，則放入 picPath array
                                if pic == picDir[y]:
                                    picPath[i - 10] = (finalFolderPath + "/" + finalDir[x])
                                    picName[i - 10] = picDir[y]

        # Progress Bar One ( 掃描檔案，建立超連結 )
        global temp1
        temp1 += 1
        print('\r' + '%.2f%% |%s%s| %d/%d' % ((float(temp1/total1*100)), '█' * int(temp1*20/total1), ' ' * (20-int(temp1*20/total1)), (i-9), (editWS.max_row-9)), end='')

    for i in range(10, editWS.max_row+1):
        HyperLinkBuilding(i)

    # ================================================================================ #
    #                                                                                  #
    #                      複製照片至 Excel 路徑底下的 Image 資料夾                      #
    #                                                                                  #
    # ================================================================================ #

    print("\n\n照片複製中...\n")

    temp2 = 0
    total2 = editWS.max_row-9

    @retry_on_communication_error
    def ImageCopy(i):
        unit = (editWS.cell(row=i, column=unitCoordinate).value)
        temp = (editWS.cell(row=i, column=tempCoordinate).value)

        if picPath[i - 10] != "Path":
            newPath = "Image/" + newImageFolder.split("/")[-1] + "/" + (picName[i - 10][:-4] + "_" + unit[-4:] + "_" + temp + ".png")
            editWS.cell(row=i, column=imageCoordinate).value = ('=HYPERLINK("%s", "Link")' % newPath)
            editWS.cell(row=i, column=imageCoordinate).font = Font(underline="single", color='00B050')
            editWS.cell(row=i, column=sessionCoordinate).value = sessionPath[i - 10]
            try:
                shutil.copyfile(os.path.join(picPath[i - 10], picName[i - 10]),os.path.join(newImageFolder, picName[i - 10][:-4] + "_" + unit[-4:] + "_" + temp + ".png"))
                #pass
            except:
                print("Directory you are copying does not exist.")
                error("第 %d 筆資料異常" % (i))
                pass

        # Progress Bar Two ( 照片複製 )
        global temp2
        temp2 += 1
        print('\r' + '%.2f%% |%s%s| %d/%d' % ((float(temp2/total2*100)), '█' * int(temp2*20/total2), ' ' * (20-int(temp2*20/total2)), (i-9), (editWS.max_row-9)), end='')

    for i in range(10, editWS.max_row+1):
        ImageCopy(i)

    # ================================================================================ #
    #                                                                                  #
    #                              檔案儲存 ( Excel, txt )                              #
    #                                                                                  #
    # ================================================================================ #

    for i in range(len(picPath)):
        if picPath[i] != "Path":
            print("\n\nSaving...  Please Wait...\n")
            print("========================================\n")
            #wb.save(excelPath.get()[:-5] + "_copy.xlsx")
            wb.save(excelPath.get())
            print("Success!\n")
            time.sleep(2)
            break
        else:
            wb.save(excelPath.get())
            print("\n")
            time.sleep(2)
            break

    f.close()

    # ================================================================================ #
    #                                                                                  #
    #                               詢問程式是否要重新執行                               #
    #                                                                                  #
    # ================================================================================ #

    print("========================================\n")
    print("Re-run Y/N ?")
    executeYN = input()
    print()
    print("========================================\n")

    if executeYN == "Y" or executeYN == "y":
        execute = True
    elif executeYN == "N" or executeYN == "n":
        execute = False
    if execute == False:
        break
